# reporter.py
import os
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


LEGEND_TEXT = """
ОБОЗНАЧЕНИЯ СИГНАЛОВ
═══════════════════════════════════

⬆ ПОКУПКА от {level}
  Цена находится у уровня поддержки.
  Рекомендуется открыть длинную позицию.
  SL = цена - ATR_SL * ATR (ниже уровня)
  TP = цена + ATR_TP * ATR

⬇ ПРОДАЖА от {level}
  Цена находится у уровня сопротивления.
  Рекомендуется открыть короткую позицию.
  SL = цена + ATR_SL * ATR (выше уровня)
  TP = цена - ATR_TP * ATR

➡ ОЖИДАНИЕ у {level}
  Цена рядом с уровнем, но сигнал
  ещё не подтверждён. Мониторить.

➡ НЕТ СИГНАЛА
  Цена далеко от всех исторических
  уровней. Позиция не рекомендуется.

Сила уровня [*****]
  [*****] = 5/5 очень сильный
  [**** ] = 4/5 сильный
  [***  ] = 3/5 средний
  [**   ] = 2/5 слабый
  [*    ] = 1/5 очень слабый

Критерии силы:
  • Кол-во касаний уровня
  • Win rate сделок на уровне
  • Свежесть касаний (последние 10)

СОКРАЩЕНИЯ В ОТЧЁТЕ
═══════════════════════════════════
  Ret — общая доходность (%)
  WR  — доля прибыльных сделок (%)
  PF  — profit factor (прибыль / убыток)

ТИП ВХОДА В СДЕЛКУ
═══════════════════════════════════
  0 — По рынку (open след. свечи)
      Сигнал на свече i → вход по open
      свечи i+1. Исполняется всегда.
  1 — По цене сигнала (лимитный)
      Вход по цене уровня на свече i+1,
      только если свеча коснулась уровня.
      Если не коснулась — пропуск.
"""


def generate_report(all_results, top_n=5, params=None):
    lines = []
    lines.append("=" * 75)
    lines.append(f"  СКАНЕР СЕКТОРОВ — {datetime.now().strftime('%d.%m.%Y %H:%M')}")
    lines.append("=" * 75)
    lines.append("")

    if not all_results:
        lines.append("  Нет результатов для отображения.")
        return '\n'.join(lines)

    if params:
        atr_sl = params.get('atr_sl', '--')
        atr_tp = params.get('atr_tp', '--')
        risk = params.get('risk_per_trade', '--')
        if isinstance(risk, float):
            risk_pct = f"{risk * 100:.1f}%"
        else:
            risk_pct = risk
        min_hits = params.get('min_hits', '--')
        capital = params.get('capital', '--')
        if isinstance(capital, (int, float)):
            capital_str = f"{capital:,.0f}"
            pos_size = capital * (risk if isinstance(risk, float) else 0.02)
            pos_size_str = f"{pos_size:,.0f} руб"
        else:
            capital_str = '--'
            pos_size_str = '--'
        last_candles = params.get('last_candles', '--')
        lines.append(
            f"  Параметры: ATR_SL={atr_sl} | ATR_TP={atr_tp} | Риск={risk_pct} | "
            f"Мин.повторов={min_hits} | Капитал={capital_str} | Риск,₽={pos_size_str} | "
            f"Свежесть={last_candles}"
        )
    else:
        lines.append(f"  Параметры: ATR_SL=-- | ATR_TP=-- | Риск=--")

    lines.append("")

    passed = sum(1 for r in all_results if r['metrics'].get('total_trades', 0) > 0)
    total = len(all_results)
    lines.append(f"  Прошло фильтр: {passed}/{total}")
    lines.append("")

    sectors_order = []
    seen_sectors = set()
    for r in all_results:
        if r['sector'] not in seen_sectors:
            sectors_order.append(r['sector'])
            seen_sectors.add(r['sector'])

    for sector in sectors_order:
        sector_results = [
            r for r in all_results
            if r['sector'] == sector and r['metrics'].get('total_trades', 0) > 0
        ]
        if not sector_results:
            continue

        lines.append(f"  --- {sector} ({len(sector_results)}) ---")

        for r in sector_results:
            m = r['metrics']
            sig = r['signal']
            action = sig.get('action', 'NONE')
            level = sig.get('level', '')
            strength = sig.get('strength', {})
            stars = strength.get('stars', '[     ]') if strength else ''
            sl_price = sig.get('sl_price')
            tp_price = sig.get('tp_price')

            action_map = {
                'BUY': '⬆ ПОКУПКА',
                'SELL': '⬇ ПРОДАЖА',
                'WAIT': '➡ ОЖИДАНИЕ',
                'NONE': '➡ НЕТ СИГНАЛА'
            }
            action_text = action_map.get(action, '➡ --')

            level_part = f"от {level}" if level and action in ('BUY', 'SELL') else ''
            wait_part = f"у {level}" if level and action == 'WAIT' else ''

            sl_part = f"SL={sl_price}" if sl_price and action in ('BUY', 'SELL') else ''
            tp_part = f"TP={tp_price}" if tp_price and action in ('BUY', 'SELL') else ''

            ret = m.get('total_return', 0)
            wr = m.get('win_rate', 0)
            pf = m.get('profit_factor', 0)
            sh = m.get('sharpe', 0)
            nt = m.get('total_trades', 0)

            ret_s = f"{ret:+.1f}%" if isinstance(ret, (int, float)) else "0%"
            wr_s = f"{wr:.0f}%" if isinstance(wr, (int, float)) else "0%"
            pf_s = f"{pf:.1f}" if isinstance(pf, (int, float)) else "0"
            sh_s = f"{sh:.1f}" if isinstance(sh, (int, float)) else "0"

            capital = params.get('capital', 1_000_000) if params else 1_000_000
            risk = params.get('risk_per_trade', 0.02) if params else 0.02
            risk_amount = capital * risk

            direction = f"{action_text} {level_part} {wait_part}".strip()
            sl_tp = f"{sl_part} {tp_part}".strip()
            if sl_tp:
                sl_tp = f" [{sl_tp}]"

            vol_text = ''
            if action in ('BUY', 'SELL') and level and sl_price:
                sl_dist = abs(float(level) - float(sl_price)) / float(level)
                if sl_dist > 0:
                    vol_text = f" {risk_amount / sl_dist:,.0f}₽"

            lines.append(
                f"  {r['ticker']:<6s} {direction:<30s} {stars:>7s}  "
                f"Ret:{ret_s:>7s}  WR:{wr_s:>4s}  PF:{pf_s:>4s}  "
                f"({nt:>2d} сд.){vol_text}{sl_tp}"
            )

        lines.append("")

    lines.append("=" * 75)
    lines.append(f"  РЕКОМЕНДАЦИИ TOP-{top_n} (по доходности)")
    lines.append("=" * 75)

    top = []
    for r in all_results:
        if r['signal']['action'] not in ('NONE',) and r['metrics'].get('total_trades', 0) >= 2:
            score = r['metrics'].get('total_return', -999)
            if r['metrics'].get('sharpe', 0) <= 0:
                continue
            top.append((score, r))
    top.sort(key=lambda x: x[0], reverse=True)
    top = top[:top_n]

    if not top:
        lines.append("  Нет бумаг, прошедших фильтр.")
    else:
        capital = params.get('capital', 1_000_000) if params else 1_000_000
        risk = params.get('risk_per_trade', 0.02) if params else 0.02
        risk_amount = capital * risk

        for i, (score, r) in enumerate(top, 1):
            m = r['metrics']
            sig = r['signal']
            action_map = {'BUY': '⬆', 'SELL': '⬇', 'WAIT': '➡'}
            arrow = action_map.get(sig.get('action', ''), '➡')
            level = sig.get('level', '')
            strength = sig.get('strength', {})
            stars = strength.get('stars', '') if strength else ''
            sl_price = sig.get('sl_price', '')
            tp_price = sig.get('tp_price', '')
            level_text = f"от {level}" if level and sig.get('action') in ('BUY', 'SELL') else ''
            sl_tp_text = f"SL={sl_price} TP={tp_price}" if sl_price and tp_price and sig.get('action') in ('BUY', 'SELL') else ''
            vol_text = ''
            if sig.get('action') in ('BUY', 'SELL') and level and sl_price:
                sl_dist = abs(float(level) - float(sl_price)) / float(level)
                if sl_dist > 0:
                    vol_text = f"Объём={risk_amount / sl_dist:,.0f}₽"
            ret = m.get('total_return', 0)
            ret_s = f"{ret:+.1f}%" if isinstance(ret, (int, float)) else "0%"

            lines.append(
                f"  {i}. {r['ticker']:<5s} {r['sector']:<18s} "
                f"{arrow} {level_text:<15s} {stars:>7s}  "
                f"Ret:{ret_s:>7s}  {vol_text}  {sl_tp_text}"
            )

    lines.append("=" * 75)
    return '\n'.join(lines)


def export_to_excel(all_results, params, filepath):
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl не установлен. pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Сканер секторов"

    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=14)
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

    row = 1
    ws.cell(row=row, column=1, value=f"Сканер секторов — {datetime.now().strftime('%d.%m.%Y %H:%M')}")
    ws.cell(row=row, column=1).font = title_font
    row += 1

    if params:
        ws.cell(row=row, column=1,
                value=f"ATR_SL={params.get('atr_sl','--')}  ATR_TP={params.get('atr_tp','--')}  "
                      f"Риск={params.get('risk_per_trade','--')}  Мин.повторов={params.get('min_hits','--')}")
        row += 1

    headers = ['Тикер', 'Сектор', 'Сигнал', 'Уровень', 'Сила', 'SL', 'TP',
               'Доходность%', 'WinRate%', 'Profit Factor', 'Sharpe', 'Сделок']
    row += 1
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font

    sectors_order = []
    seen = set()
    for r in all_results:
        if r['sector'] not in seen:
            sectors_order.append(r['sector'])
            seen.add(r['sector'])

    row += 1
    for sector in sectors_order:
        sector_results = [
            r for r in all_results
            if r['sector'] == sector and r['metrics'].get('total_trades', 0) > 0
        ]
        if not sector_results:
            continue

        ws.cell(row=row, column=1, value=sector).font = Font(bold=True, size=11)
        row += 1

        for r in sector_results:
            m = r['metrics']
            sig = r['signal']
            action_map = {'BUY': 'ПОКУПКА', 'SELL': 'ПРОДАЖА', 'WAIT': 'ОЖИДАНИЕ', 'NONE': 'НЕТ СИГНАЛА'}
            ws.cell(row=row, column=1, value=r['ticker'])
            ws.cell(row=row, column=2, value=sector)
            ws.cell(row=row, column=3, value=action_map.get(sig.get('action', ''), ''))
            ws.cell(row=row, column=4, value=sig.get('level', ''))
            ws.cell(row=row, column=5, value=sig.get('strength', {}).get('strength', ''))
            ws.cell(row=row, column=6, value=sig.get('sl_price', ''))
            ws.cell(row=row, column=7, value=sig.get('tp_price', ''))
            ws.cell(row=row, column=8, value=round(m.get('total_return', 0), 1))
            ws.cell(row=row, column=9, value=round(m.get('win_rate', 0), 1))
            ws.cell(row=row, column=10, value=round(m.get('profit_factor', 0), 2))
            ws.cell(row=row, column=11, value=round(m.get('sharpe', 0), 2))
            ws.cell(row=row, column=12, value=m.get('total_trades', 0))

            ret_cell = ws.cell(row=row, column=8)
            if m.get('total_return', 0) > 0:
                ret_cell.fill = green_fill
            else:
                ret_cell.fill = red_fill

            row += 1

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 14
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 14
    ws.column_dimensions['K'].width = 10
    ws.column_dimensions['L'].width = 10

    wb.save(filepath)
    return filepath


def export_smart_scan_excel(all_results, params, filepath):
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl не установлен. pip install openpyxl")

    wb = openpyxl.Workbook()

    # --- Лист 1: Сводка ---
    ws = wb.active
    ws.title = "Лучшая стратегия"
    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=14)
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

    from strategy.config import get_strategy_names
    strategy_reverse = {sid: name for sid, name in get_strategy_names()}

    row = 1
    ws.cell(row=row, column=1, value=f"Умный сканер — {datetime.now().strftime('%d.%m.%Y %H:%M')}")
    ws.cell(row=row, column=1).font = title_font
    row += 1

    if params:
        ws.cell(row=row, column=1,
                value=f"ATR_SL={params.get('atr_sl','--')}  ATR_TP={params.get('atr_tp','--')}  "
                      f"Риск={params.get('risk_per_trade','--')}  Мин.повторов={params.get('min_hits','--')}")
        row += 1

    headers = ['№', 'Тикер', 'Сектор', 'Лучшая стратегия',
               'Доходность%', 'Sharpe', 'Сделок', 'Сигнал']
    row += 1
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font

    for rank, r in enumerate(all_results, 1):
        row += 1
        best_sid = r.get('best_strategy')
        best_name = strategy_reverse.get(best_sid, best_sid or '—')
        metrics = r.get('best_metrics', {})
        sig = r.get('best_signal', {})
        action_map = {'BUY': '⬆ ПОКУПКА', 'SELL': '⬇ ПРОДАЖА', 'WAIT': '➡ ОЖИДАНИЕ', 'NONE': '—'}
        ret = metrics.get('total_return', 0)
        sh = metrics.get('sharpe', 0)
        tr = metrics.get('total_trades', 0)

        ws.cell(row=row, column=1, value=rank)
        ws.cell(row=row, column=2, value=r['ticker'])
        ws.cell(row=row, column=3, value=r['sector'])
        ws.cell(row=row, column=4, value=best_name if best_sid else '—')
        ws.cell(row=row, column=5, value=round(ret, 1) if isinstance(ret, (int, float)) else '—')
        ws.cell(row=row, column=6, value=round(sh, 2) if isinstance(sh, (int, float)) else '—')
        ws.cell(row=row, column=7, value=tr if tr else '—')
        ws.cell(row=row, column=8, value=action_map.get(sig.get('action', ''), ''))

        ret_cell = ws.cell(row=row, column=5)
        if isinstance(ret, (int, float)):
            ret_cell.fill = green_fill if ret > 0 else red_fill

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 8
    ws.column_dimensions['H'].width = 14

    # --- Лист 2: Матрица тикер × стратегия ---
    ws2 = wb.create_sheet("Матрица стратегий")
    strategy_names_ordered = [(sid, name) for sid, name in get_strategy_names()]

    row = 1
    ws2.cell(row=row, column=1, value="Умный сканер — матрица стратегий").font = title_font
    row += 1

    ws2.cell(row=row, column=2, value="Best")
    ws2.cell(row=row, column=2).font = header_font
    for ci, (sid, name) in enumerate(strategy_names_ordered, 3):
        ws2.cell(row=row, column=ci, value=name).font = header_font
    row += 1

    for r in all_results:
        strategies = r.get('strategies', {})
        best_sid = r.get('best_strategy')

        ws2.cell(row=row, column=1, value=r['ticker'])
        ws2.cell(row=row, column=2, value=strategy_reverse.get(best_sid, '—') if best_sid else '—')

        for ci, (sid, name) in enumerate(strategy_names_ordered, 3):
            sdata = strategies.get(sid)
            if sdata:
                metrics = sdata.get('metrics', {})
                ret = metrics.get('total_return', 0)
                ret_s = f"{ret:+.1f}%" if isinstance(ret, (int, float)) else '—'
                ws2.cell(row=row, column=ci, value=ret_s)
                cell = ws2.cell(row=row, column=ci)
                if isinstance(ret, (int, float)):
                    cell.fill = green_fill if ret > 0 else red_fill
            else:
                ws2.cell(row=row, column=ci, value='—')
        row += 1

    ws2.column_dimensions['A'].width = 8
    ws2.column_dimensions['B'].width = 18
    for ci in range(3, 3 + len(strategy_names_ordered)):
        ws2.column_dimensions[chr(64 + ci) if ci <= 26 else 'A'].width = 16

    wb.save(filepath)
    return filepath


def get_legend_text():
    return LEGEND_TEXT
