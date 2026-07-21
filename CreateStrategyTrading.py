# CreateStrategyTrading.py
# -*- coding: utf-8 -*-
import json
import logging
import os
import sys
import threading
import traceback
import tempfile
from datetime import datetime, timedelta

import pandas as pd
import requests
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox as mb
try:
    import sv_ttk
    HAS_SV_TTK = True
except ImportError:
    HAS_SV_TTK = False

from core.moex_session import MOEX_SESSION as _MOEX_SESSION

from backtest.engine import BacktestEngine, export_results as _export_results, candles_to_df
from screening.levels_strength import DEFAULT_LAST_CANDLES, calculate_level_strength, get_best_level_signal
from screening.reporter import generate_report
from screening.scanner import Scanner
from screening.smart_scanner import SmartScanner
from screening.sectors import SectorDB
from strategy.bounce import check_bounce
from strategy.indicators import calc_atr
from strategy.levels import find_strong_zones
from visual import StockAppVisual, ScannerUI, SmartScannerUI, DiaryUI, StrategyGuideUI, AppGuideUI, _add_copy_menu
from diary.journal import DiaryStorage, DiaryEntry, calc_position_qty, calc_position_volume
from intraday.visual import IntradayUI
from automation.scheduler import AutomationScheduler
from automation.panel import AutomationPanel
from cloud.ui import CloudPanel
from settings_dialog import SettingsDialog
from utils import normalize_numeric_params, migrate_ticker_settings, load_favorites, toggle_favorite, sort_tickers_by_favorites, app_dir, tree_batch_insert
from core.moex_cache import moex_cache, cached_get_tickers
from core.session_store import get_cached_range, save_session, merge_candles
from pending.storage import PendingTradesStorage
from pending.monitor import PendingTradesMonitor
from pending.ui import PendingTradesUI

# Настройка логирования
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(os.path.join(app_dir(), 'CreateStrategyTrading.log'), encoding='utf-8'),
        logging.StreamHandler()
    ]
)

# Константы
MAX_DATA_FETCH_INTERVAL = timedelta(days=365 * 2)
DEFAULT_START_DATE = "2020-01-01"
DEFAULT_END_DATE = datetime.now().strftime("%Y-%m-%d")
MIN_CANDLES_FOR_BACKTEST = 30
INITIAL_CAPITAL = 1_000_000
DEFAULT_MIN_REPEATS = 5
SESSION_STATE_PATH = os.path.join(app_dir(), 'results', 'session_state.json')


class StrategyAppState:
    """Хранение состояния приложения"""
    def __init__(self):
        self.stock_data = None
        self.current_step = 0
        self.strong_zones = []
        self.df = None


class CreateStrategyApp:
    """Основное приложение для анализа и backtesting стратегий"""
    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.root.withdraw()

        splash = tk.Toplevel(root)
        splash.title("CreateStrategy")
        splash.resizable(False, False)
        splash.overrideredirect(True)
        sw = root.winfo_screenwidth()
        sh = root.winfo_screenheight()
        splash.geometry(f"420x180+{(sw-420)//2}+{(sh-180)//2}")

        _frm = tk.Frame(splash, bd=2, relief=tk.GROOVE)
        _frm.pack(fill=tk.BOTH, expand=True)
        tk.Label(_frm, text="CreateStrategy", font=("Segoe UI", 16, "bold")).pack(pady=(15, 2))
        _lbl = tk.Label(_frm, text="Инициализация...", font=("Segoe UI", 10))
        _lbl.pack()
        _pb = ttk.Progressbar(_frm, mode="indeterminate", length=360)
        _pb.pack(pady=8)
        _pb.start(15)
        splash.update()

        self.state = StrategyAppState()
        self.app = None
        self.scanner_ui = None
        self.diary_ui = None
        self.watchlist_ui = None

        _lbl.config(text="Загрузка секторов...")
        splash.update()
        self.sector_db = SectorDB()

        self._last_scan_results = []
        self._last_scan_params = {}
        self._last_smart_results = []
        self._last_smart_params = {}
        self._last_capital = 1_000_000
        self._last_trades = None
        self._last_metrics = None
        self._last_export_stock = None
        self._favorites = load_favorites()
        self._current_regime = None

        _lbl.config(text="Загрузка дневника...")
        splash.update()
        self.diary_storage = DiaryStorage()

        _lbl.config(text="Загрузка уведомлений...")
        splash.update()
        from monitoring.notification_manager import NotificationManager
        self.notification_manager = NotificationManager()

        migrate_ticker_settings(os.path.join(app_dir(), 'results', 'ticker_settings.json'))

        _lbl.config(text="Построение интерфейса...")
        splash.update()
        self.setup_ui()
        self._load_session_state()
        self._start_auto_load_tickers()
        self.root.protocol('WM_DELETE_WINDOW', self._on_close)
        self.bind_events()
        self.root.after(1000, self._auto_check_positions)
        self.root.after(3000, self._update_market_regime)
        self.root.after(2000, self._init_new_tabs)

        _lbl.config(text="Запуск автоматизации...")
        splash.update()
        self._init_automation()

        _pb.stop()
        splash.destroy()
        self.root.deiconify()

    def candles_to_df_custom(self, candles_list: list) -> pd.DataFrame | None:
        """Конвертация списка свечей в DataFrame"""
        valid = [c for c in candles_list if c is not None and len(c) > 6]
        if not valid:
            return None
        return pd.DataFrame(
            valid,
            columns=['Open', 'Close', 'High', 'Low', 'Volume', 'Value', 'Begin', 'End'],
            index=pd.to_datetime([c[6] for c in valid], format='mixed')
        )

    def get_stock_data(self, symbol: str, start_date: str, end_date: str) -> list | str:
        """Получение исторических данных по акции (с инкрементальным кэшем)"""
        cached = get_cached_range(symbol, 24, start_date, end_date)
        if cached is not None:
            return cached

        from core.session_store import load_session
        session = load_session(symbol, 24)
        if session is not None:
            s_last = session.get('last_date', '')
            s_start = session.get('start_date', '')
            try:
                last_dt = datetime.strptime(s_last, '%Y-%m-%d')
                end_dt = datetime.strptime(end_date, '%Y-%m-%d')
                start_dt = datetime.strptime(start_date, '%Y-%m-%d')
                c_start_dt = datetime.strptime(s_start, '%Y-%m-%d') if s_start else None

                need_tail = last_dt < end_dt
                need_head = c_start_dt and c_start_dt > start_dt

                if need_tail or need_head:
                    result_candles = list(session['candles'])

                    if need_tail:
                        tail_start = max(s_last, start_date)
                        tail = self._fetch_stock_data_raw(symbol, tail_start, end_date)
                        if isinstance(tail, list) and tail:
                            result_candles = merge_candles(result_candles, tail)

                    if need_head:
                        head_end = min(s_start, end_date)
                        head = self._fetch_stock_data_raw(symbol, start_date, head_end)
                        if isinstance(head, list) and head:
                            result_candles = merge_candles(head, result_candles)

                    save_session(symbol, 24, result_candles, start_date=start_date)
                    return result_candles

                return result_candles
            except ValueError:
                pass

        result = self._fetch_stock_data_raw(symbol, start_date, end_date)
        if isinstance(result, list) and result:
            save_session(symbol, 24, result, start_date=start_date)
        return result

    def _fetch_stock_data_raw(self, symbol: str, start_date: str, end_date: str) -> list | str:
        """Получение исторических данных по акции (сырой HTTP-запрос)"""
        try:
            url = f"https://iss.moex.com/iss/engines/stock/markets/shares/boards/TQBR/securities/{symbol}/candles.json"
            start_dt = datetime.strptime(start_date, "%Y-%m-%d")
            end_dt = datetime.strptime(end_date, "%Y-%m-%d")
            all_candles = []
            current_start = start_dt

            while current_start < end_dt:
                current_end = min(current_start + MAX_DATA_FETCH_INTERVAL, end_dt)
                params = {
                    "start": 0,
                    "till": current_end.strftime("%Y-%m-%d"),
                    "from": current_start.strftime("%Y-%m-%d"),
                    "interval": 24
                }
                response = _MOEX_SESSION.get(url, params=params, timeout=30)
                response.raise_for_status()
                data = response.json()
                if "candles" in data and "data" in data["candles"]:
                    all_candles.extend(data["candles"]["data"])
                current_start += MAX_DATA_FETCH_INTERVAL
            return all_candles if all_candles else "Нет данных за указанный период"
        except requests.exceptions.HTTPError as e:
            logging.error(f"Ошибка HTTP: {e}")
            return f"Ошибка HTTP: {e}"
        except Exception as e:
            logging.error(f"Ошибка получения данных: {e}")
            return f"Ошибка получения данных: {e}"

    DARK_TEXT_BG = '#1e1e1e'
    DARK_TEXT_FG = '#d4d4d4'
    LIGHT_TEXT_BG = '#ffffff'
    LIGHT_TEXT_FG = '#000000'

    def setup_ui(self) -> None:
        """Настройка GUI"""
        self.root.title("CreateStrategy — Технический анализ MOEX")
        self.root.geometry("1250x800")

        self._current_theme = 'light'
        if HAS_SV_TTK:
            sv_ttk.set_theme(self._current_theme)

        top_bar = ttk.Frame(self.root)
        top_bar.pack(fill=tk.X, padx=5, pady=(5, 0))
        ttk.Label(top_bar, text="CreateStrategy", font=('', 12, 'bold')).pack(side=tk.LEFT)

        self.regime_label = ttk.Label(top_bar, text="Рынок: ...", font=('', 10), cursor='hand2')
        self.regime_label.pack(side=tk.LEFT, padx=20)
        self.regime_label.bind('<Button-1>', self._on_regime_click)

        self.breadth_label = ttk.Label(top_bar, text="Breadth: ...", font=('', 10), cursor='hand2')
        self.breadth_label.pack(side=tk.LEFT, padx=5)
        self.breadth_label.bind('<Button-1>', self._on_breadth_click)

        if HAS_SV_TTK:
            self._theme_btn = ttk.Button(top_bar, text='🌙 Тёмная', width=12, command=self._toggle_theme)
            self._theme_btn.pack(side=tk.RIGHT)

        self._notif_btn = ttk.Button(top_bar, text='🔔 Уведомления', width=16,
                                     command=self._show_notification_settings)
        self._notif_btn.pack(side=tk.RIGHT, padx=5)

        self._settings_btn = ttk.Button(top_bar, text='Настройки', width=12,
                                         command=self._show_settings)
        self._settings_btn.pack(side=tk.RIGHT, padx=(0, 5))

        notebook = ttk.Notebook(self.root)
        notebook.pack(fill='both', expand=1)

        tab_analysis = ttk.Frame(notebook)
        tab_scanner = ttk.Frame(notebook)
        tab_smart_scanner = ttk.Frame(notebook)
        tab_diary = ttk.Frame(notebook)
        tab_pending = ttk.Frame(notebook)
        tab_positions = ttk.Frame(notebook)
        tab_review = ttk.Frame(notebook)
        tab_analytics = ttk.Frame(notebook)
        tab_pairs = ttk.Frame(notebook)
        tab_sectors = ttk.Frame(notebook)
        tab_calc = ttk.Frame(notebook)
        tab_signals = ttk.Frame(notebook)
        tab_alerts = ttk.Frame(notebook)
        tab_news = ttk.Frame(notebook)
        tab_app_guide = ttk.Frame(notebook)
        notebook.add(tab_analysis, text='Анализ')
        notebook.add(tab_scanner, text='Сканер')
        notebook.add(tab_smart_scanner, text='Умный сканер')
        notebook.add(tab_diary, text='Дневник сделок')
        notebook.add(tab_pending, text='Ожидание')
        notebook.add(tab_positions, text='Позиции')
        notebook.add(tab_review, text='Обзор торговли')
        notebook.add(tab_analytics, text='Аналитика')
        notebook.add(tab_pairs, text='Пары')
        notebook.add(tab_sectors, text='Секторы')
        notebook.add(tab_calc, text='Калькулятор')
        notebook.add(tab_signals, text='Сигналы')
        notebook.add(tab_alerts, text='Алерты')
        notebook.add(tab_news, text='AI Новости')
        notebook.add(tab_app_guide, text='Описание')

        tab_guide = ttk.Frame(notebook)
        notebook.add(tab_guide, text='Справочник стратегий')

        tab_intraday = ttk.Frame(notebook)
        notebook.add(tab_intraday, text='Интрадей')

        intraday_notebook = ttk.Notebook(tab_intraday)
        intraday_notebook.pack(fill='both', expand=1)

        tab_intra_backtest = ttk.Frame(intraday_notebook)
        tab_intra_scanner = ttk.Frame(intraday_notebook)
        intraday_notebook.add(tab_intra_backtest, text='Backtest')
        intraday_notebook.add(tab_intra_scanner, text='Умный сканер')

        self.app = StockAppVisual(
            tab_analysis, self.on_select, self.on_export_button,
            self.get_moex_tickers, self.on_backtest,
            on_diary=self.on_add_to_diary_analysis,
            on_show_settings=self.on_show_ticker_settings,
            on_save_results=self._on_save_results,
            on_optimize=self.on_optimize,
            on_portfolio=self.on_portfolio,
            on_walkforward=self.on_walkforward,
            on_sensitivity=self.on_sensitivity,
            on_fundamental=self._on_fundamental,
            on_heatmap=self.on_heatmap,
            favorites=self._favorites,
            on_toggle_favorite=self._on_toggle_favorite,
            sector_db=self.sector_db,
            on_pending=self.on_add_to_pending,
        )

        all_sectors = self.sector_db.get_all_sectors()
        total_tickers = len(self.sector_db.get_tickers(all_sectors))
        self.scanner_ui = ScannerUI(
            tab_scanner, sectors=all_sectors,
            on_scan=self.on_scanner, on_legend=self.on_show_legend,
            on_excel=self.on_export_excel, on_diary=self.on_add_to_diary,
            on_show_settings=self.on_show_ticker_settings,
            total_tickers=total_tickers
        )

        self.smart_scanner_ui = SmartScannerUI(
            tab_smart_scanner, sectors=all_sectors,
            on_scan=self.on_smart_scanner,
            on_excel=self.on_smart_export_excel,
            on_diary=self.on_add_to_diary,
            total_tickers=total_tickers
        )

        self.diary_ui = DiaryUI(
            tab_diary, storage=self.diary_storage,
            on_check_positions=self.on_check_positions,
            on_show_analysis=self.on_show_analysis
        )

        self.pending_storage = PendingTradesStorage()
        self.pending_ui = PendingTradesUI(
            tab_pending,
            on_remove=self._on_pending_remove,
            on_refresh=self._on_pending_refresh,
        )
        self.pending_ui.update_trades(self.pending_storage.get_all())
        self.pending_monitor = None

        from visual import PositionDashboardUI
        self.position_ui = PositionDashboardUI(
            tab_positions,
            on_refresh=self._on_position_refresh,
            on_start_monitor=self._on_position_start,
            on_stop_monitor=self._on_position_stop,
        )
        self.position_monitor = None

        from visual import TradeReviewUI
        self.review_ui = TradeReviewUI(
            tab_review,
            on_refresh=self._on_review_refresh,
        )

        from visual import PerformanceAnalyticsUI
        self.perf_ui = PerformanceAnalyticsUI(
            tab_analytics,
            on_analyze=self._on_perf_analyze,
        )

        from visual import PairsTradingUI
        self.pairs_ui = PairsTradingUI(
            tab_pairs,
            on_scan=self._on_pairs_scan,
            on_backtest=self._on_pairs_backtest,
        )

        from visual import WatchlistUI, SectorRotationUI, PositionCalculatorUI, SignalJournalUI
        from watchlist.storage import WatchlistStorage
        from signals.storage import SignalStorage

        self.watchlist_storage = WatchlistStorage()
        _all_tickers = self.sector_db.get_all_tickers()
        from strategy.config import get_strategy_names
        _all_strategy_ids = [s[0] for s in get_strategy_names()]

        self._watchlist_ui_class = WatchlistUI
        self._watchlist_callbacks = {
            'on_add': self._on_watchlist_add,
            'on_remove': self._on_watchlist_remove,
            'on_refresh': self._on_watchlist_refresh,
            'on_select': self._on_watchlist_select,
            'on_dividends': self._on_dividend_calendar,
            'on_correlation': self._on_correlation_matrix,
            'all_tickers': _all_tickers,
        }

        self.sector_rotation_ui = SectorRotationUI(
            tab_sectors,
            on_scan=self._on_sector_scan,
            on_compare=self._on_sector_compare,
        )

        self.pos_calc_ui = PositionCalculatorUI(
            tab_calc,
            on_calculate=self._on_pos_calc,
            get_current_price=self._get_current_price,
            all_tickers=_all_tickers,
        )

        self.signal_storage = SignalStorage()
        self.signal_journal_ui = SignalJournalUI(
            tab_signals,
            on_filter=self._on_signal_filter,
            on_export=self._on_signal_export,
            all_tickers=_all_tickers,
            all_strategies=_all_strategy_ids,
        )

        self.guide_ui = StrategyGuideUI(tab_guide)
        self.app_guide_ui = AppGuideUI(tab_app_guide)

        from alerts.storage import AlertStorage
        from alerts.monitor import AlertMonitor
        from alerts.ui import AlertUI
        self.alert_storage = AlertStorage()
        self.alert_monitor = AlertMonitor(
            self.root, self.alert_storage, self._get_current_price,
            on_triggered=self._on_alert_triggered,
            on_refresh=self._on_alert_refresh,
        )
        self.alert_ui = AlertUI(
            tab_alerts, self.alert_storage,
            on_add=self._on_alert_add,
            on_remove=self._on_alert_remove,
            on_start=self._on_alert_start,
            on_stop=self._on_alert_stop,
            all_tickers=_all_tickers,
        )

        from news.analyzer import NewsAnalyzer
        from news.ui import NewsUI
        self.news_analyzer = NewsAnalyzer(known_tickers=_all_tickers)
        self.news_ui = NewsUI(
            tab_news,
            on_refresh=self._on_news_refresh,
            on_analyze=self._on_news_analyze,
            on_settings=self._on_news_settings,
            all_tickers=_all_tickers,
        )
        cached = self.news_analyzer.get_cached()
        if cached:
            self.news_ui.update_news(cached)

        self.scheduler = AutomationScheduler(self.root)
        self.scheduler.register_task('auto_scan', self._auto_scan_callback)
        self.scheduler.register_task('monitor_positions', self._auto_monitor_callback)
        self.scheduler.register_task('refresh_watchlist', self._auto_watchlist_callback)
        self.scheduler.register_task('auto_news_scan', self._auto_news_callback)
        self.scheduler.register_task('check_pending_trades', self._auto_pending_check_callback)

        self.settings_dialog = SettingsDialog(
            self.root,
            scheduler=self.scheduler,
            on_start_all=self._on_automation_start_all,
            on_stop_all=self._on_automation_stop_all,
            watchlist_ui_class=self._watchlist_ui_class,
            watchlist_callbacks=self._watchlist_callbacks,
            main_app=self,
        )

        from intraday.strategies import SOLABUTO_REGISTRY
        intraday_tickers = self.sector_db.get_tickers(self.sector_db.get_all_sectors())
        self.intraday_ui = IntradayUI(
            tab_intra_backtest,
            on_diary_entry=self._on_intraday_diary_entry,
            fetch_fn=None
        )
        self.intraday_ui.set_tickers(intraday_tickers)

        from intraday.visual import IntradaySmartScannerUI
        self.intraday_smart_ui = IntradaySmartScannerUI(
            tab_intra_scanner,
            on_scan=self._on_intraday_smart_scan,
            on_excel=self._on_intraday_smart_excel,
            on_diary=self._on_intraday_smart_diary,
        )
        self.intraday_smart_ui.set_tickers(intraday_tickers)

    def _on_intraday_diary_entry(self, ticker, signal, params):
        side_map = {'BUY': 'LONG', 'SELL': 'SHORT'}
        entry_price = signal.get('entry_price', 0) or signal.get('level', 0)
        sl_price = signal.get('sl_price', 0)
        tp_price = signal.get('tp_price', 0)
        capital = params.get('capital', 1_000_000)
        risk_per_trade = params.get('risk_per_trade', 0.02)
        qty = calc_position_qty(capital, risk_per_trade, entry_price, sl_price)
        volume = calc_position_volume(capital, risk_per_trade, entry_price, sl_price)
        entry = DiaryEntry(
            date=datetime.now().strftime('%Y-%m-%d %H:%M'),
            ticker=ticker,
            side=side_map.get(signal['action'], signal['action']),
            entry_price=entry_price,
            sl_price=sl_price,
            tp_price=tp_price,
            volume=volume,
            qty=qty,
            max_hold=params.get('max_hold', 20),
            status='open'
        )
        self.diary_storage.add_entries([entry])
        if self.diary_ui:
            self.diary_ui.refresh()
        mb.showinfo('В дневник', f'{ticker} (H1) добавлен в дневник.')

    def _on_intraday_smart_scan(self) -> None:
        from intraday.smart_scanner import IntradaySmartScanner
        params = self.intraday_smart_ui.get_backtest_params()
        if params is None:
            self.intraday_smart_ui.show_results([])
            return
        date_from = self.intraday_smart_ui.date_from.get()
        date_to = self.intraday_smart_ui.date_to.get()
        tickers = self.intraday_smart_ui.get_scan_tickers()
        if not tickers or not date_from or not date_to:
            self.intraday_smart_ui.show_results([])
            return
        min_trades = 10
        self._last_intra_smart_params = params
        def run():
            try:
                scanner = IntradaySmartScanner()
                results = scanner.scan(
                    tickers=tickers,
                    date_from=date_from,
                    date_to=date_to,
                    base_params=params,
                    min_trades=min_trades,
                    progress_fn=lambda c, t, tk, sn: self.intraday_smart_ui.update_progress(c, t, tk, sn),
                )
                self._last_intra_smart_results = results
                self.root.after(0, lambda: self.intraday_smart_ui.show_results(results))
            except Exception as e:
                self.root.after(0, lambda: self.intraday_smart_ui.status_var.set(f'Ошибка: {str(e)}'))
            finally:
                self.root.after(0, lambda: self.intraday_smart_ui.set_running(False))
        self.intraday_smart_ui.set_running(True)
        t = threading.Thread(target=run, daemon=True)
        t.start()

    def _on_intraday_smart_excel(self) -> None:
        results = getattr(self, '_last_intra_smart_results', [])
        if not results:
            self.intraday_smart_ui.status_var.set('Ошибка: сначала запустите сканер.')
            return
        import os
        os.makedirs(os.path.join(app_dir(), 'results'), exist_ok=True)
        from datetime import datetime
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        fpath = os.path.join(app_dir(), f'results/intraday_smart_scanner_{ts}.xlsx')
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            names = {k: v['name'] for k, v in __import__('intraday.strategies', fromlist=['SOLABUTO_REGISTRY']).SOLABUTO_REGISTRY.items()}
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Intraday Smart Scan'
            headers = ['№', 'Тикер', 'Лучшая стратегия', 'Score', 'Доходность',
                       'Sharpe', 'Сделок', 'WinRate', 'Profit Factor', 'Сигнал']
            for ci, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=ci, value=h)
                cell.font = Font(bold=True)
            for ri, r in enumerate(results, 2):
                sid = r.get('best_strategy', '')
                m = r.get('best_metrics', {})
                sig = r.get('best_signal', {})
                ws.cell(row=ri, column=1, value=ri - 1)
                ws.cell(row=ri, column=2, value=r['ticker'])
                ws.cell(row=ri, column=3, value=names.get(sid, sid))
                ws.cell(row=ri, column=4, value=round(r.get('best_score', -1), 2))
                ws.cell(row=ri, column=5, value=round(m.get('total_return', 0), 2))
                ws.cell(row=ri, column=6, value=round(m.get('sharpe', 0), 2))
                ws.cell(row=ri, column=7, value=m.get('total_trades', 0))
                ws.cell(row=ri, column=8, value=round(m.get('win_rate', 0), 1))
                ws.cell(row=ri, column=9, value=round(m.get('profit_factor', 0), 2))
                ws.cell(row=ri, column=10, value=sig.get('action', 'NONE'))
            ws.column_dimensions['A'].width = 5
            ws.column_dimensions['B'].width = 10
            ws.column_dimensions['C'].width = 22
            ws.column_dimensions['D'].width = 8
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 8
            ws.column_dimensions['G'].width = 8
            ws.column_dimensions['H'].width = 8
            ws.column_dimensions['I'].width = 10
            ws.column_dimensions['J'].width = 8
            wb.save(fpath)
            self.intraday_smart_ui.status_var.set(f'Excel сохранён: {fpath}')
        except ImportError:
            self.intraday_smart_ui.status_var.set('Ошибка: установите openpyxl (pip install openpyxl)')
        except Exception as e:
            self.intraday_smart_ui.status_var.set(f'Ошибка экспорта: {str(e)}')

    def _on_intraday_smart_diary(self) -> None:
        results = getattr(self, '_last_intra_smart_results', [])
        if not results:
            self.intraday_smart_ui.status_var.set('Ошибка: сначала запустите сканер.')
            return
        params = getattr(self, '_last_intra_smart_params', {})
        capital = params.get('capital', 1_000_000)
        risk_per_trade = params.get('risk_per_trade', 0.02)
        max_hold = params.get('max_hold', 20)
        actionable = [r for r in results if r.get('best_signal', {}).get('action') in ('BUY', 'SELL')]
        if not actionable:
            self.intraday_smart_ui.status_var.set('Нет сигналов BUY/SELL для добавления.')
            return
        win = tk.Toplevel(self.root)
        win.title('Добавить в дневник (H1 Smart Scan)')
        win.geometry('750x500')
        win.transient(self.root)
        win.grab_set()
        frame = ttk.Frame(win)
        frame.pack(fill=tk.BOTH, expand=1, padx=10, pady=10)
        vars_map = {}
        list_frame = ttk.Frame(frame)
        list_frame.pack(fill=tk.BOTH, expand=1)
        canvas = tk.Canvas(list_frame)
        scrollbar = ttk.Scrollbar(list_frame, orient='vertical', command=canvas.yview)
        scrollable = ttk.Frame(canvas)
        scrollable.bind('<Configure>', lambda e: canvas.configure(scrollregion=canvas.bbox('all')))
        canvas.create_window((0, 0), window=scrollable, anchor='nw')
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=1)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        ttk.Label(scrollable, text='', font=('', 9, 'bold')).grid(row=0, column=0, sticky='w', padx=2)
        ttk.Label(scrollable, text='Тикер', font=('', 9, 'bold')).grid(row=0, column=1, sticky='w', padx=5)
        ttk.Label(scrollable, text='Сигнал', font=('', 9, 'bold')).grid(row=0, column=2, sticky='w', padx=5)
        ttk.Label(scrollable, text='Цена', font=('', 9, 'bold')).grid(row=0, column=3, sticky='w', padx=5)
        ttk.Label(scrollable, text='SL', font=('', 9, 'bold')).grid(row=0, column=4, sticky='w', padx=5)
        ttk.Label(scrollable, text='TP', font=('', 9, 'bold')).grid(row=0, column=5, sticky='w', padx=5)
        ttk.Label(scrollable, text='Объём (₽)', font=('', 9, 'bold')).grid(row=0, column=6, sticky='w', padx=5)
        for idx, r in enumerate(actionable, 1):
            sig = r['best_signal']
            ticker = r['ticker']
            action = sig['action']
            entry_price = sig.get('last_price') or sig.get('level') or 0
            sl_price = sig.get('sl_price', 0)
            tp_price = sig.get('tp_price', 0)
            qty = calc_position_qty(capital, risk_per_trade, entry_price, sl_price)
            volume = calc_position_volume(capital, risk_per_trade, entry_price, sl_price)
            var = tk.BooleanVar(value=False)
            vars_map[ticker] = {
                'var': var, 'ticker': ticker, 'action': action,
                'entry_price': entry_price, 'sl_price': sl_price, 'tp_price': tp_price,
                'qty': qty, 'volume': volume,
                'date': datetime.now().strftime('%Y-%m-%d %H:%M'),
            }
            ttk.Checkbutton(scrollable, variable=var).grid(row=idx, column=0, padx=2, pady=1)
            ttk.Label(scrollable, text=ticker).grid(row=idx, column=1, sticky='w', padx=5)
            ttk.Label(scrollable, text=action).grid(row=idx, column=2, sticky='w', padx=5)
            ttk.Label(scrollable, text=f'{entry_price:.2f}').grid(row=idx, column=3, sticky='w', padx=5)
            ttk.Label(scrollable, text=f'{sl_price:.2f}' if sl_price else '-').grid(row=idx, column=4, sticky='w', padx=5)
            ttk.Label(scrollable, text=f'{tp_price:.2f}' if tp_price else '-').grid(row=idx, column=5, sticky='w', padx=5)
            ttk.Label(scrollable, text=f'{volume:,.0f}').grid(row=idx, column=6, sticky='w', padx=5)
        btn_frame = ttk.Frame(frame)
        btn_frame.pack(fill=tk.X, pady=(10, 0))
        def add_selected():
            entries = []
            for info in vars_map.values():
                if info['var'].get():
                    side_map = {'BUY': 'LONG', 'SELL': 'SHORT'}
                    entries.append(DiaryEntry(
                        date=info['date'], ticker=info['ticker'],
                        side=side_map.get(info['action'], info['action']),
                        entry_price=info['entry_price'], sl_price=info['sl_price'],
                        tp_price=info['tp_price'], volume=info['volume'], qty=info['qty'],
                        max_hold=max_hold, status='open',
                    ))
            if entries:
                self.diary_storage.add_entries(entries)
                if self.diary_ui:
                    self.diary_ui.refresh()
                win.destroy()
                self.intraday_smart_ui.status_var.set(f'Добавлено в дневник: {len(entries)} сделок.')
            else:
                mb.showwarning('Нет выбора', 'Не выбрано ни одной сделки.')
        ttk.Button(btn_frame, text='Добавить выбранные', command=add_selected).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text='Отмена', command=win.destroy).pack(side=tk.LEFT, padx=5)
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), 'units')
        canvas.bind_all('<MouseWheel>', _on_mousewheel)
        win.protocol('WM_DELETE_WINDOW', lambda: (canvas.unbind_all('<MouseWheel>'), win.destroy()))

    def bind_events(self) -> None:
        """Привязка событий"""
        pass

    def _save_session_state(self) -> None:
        import json
        os.makedirs(os.path.join(app_dir(), 'results'), exist_ok=True)
        state = {
            'last_ticker': self.app.get_selected_ticker(),
            'last_capital': self._last_capital,
            'theme': self._current_theme,
        }
        try:
            with open(SESSION_STATE_PATH, 'w', encoding='utf-8') as f:
                json.dump(state, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    def _load_session_state(self) -> None:
        import json
        if not os.path.exists(SESSION_STATE_PATH):
            return
        try:
            with open(SESSION_STATE_PATH, 'r', encoding='utf-8') as f:
                state = json.load(f)
            if 'last_ticker' in state and state['last_ticker']:
                self.app.stock_combobox.set(state['last_ticker'])
                self.app._load_ticker_settings()
            if 'last_capital' in state:
                self._last_capital = state['last_capital']
            if 'theme' in state and state['theme'] == 'dark':
                self._toggle_theme()
        except Exception:
            pass

    def _toggle_theme(self):
        if not HAS_SV_TTK:
            return
        self._current_theme = 'dark' if self._current_theme == 'light' else 'light'
        sv_ttk.set_theme(self._current_theme)
        self._apply_text_theme()
        if self._theme_btn:
            self._theme_btn.config(text='☀️ Светлая' if self._current_theme == 'dark' else '🌙 Тёмная')

    def _show_notification_settings(self):
        from visual import NotificationSettingsUI
        NotificationSettingsUI(self.root, self.notification_manager)

    def _show_settings(self):
        self.settings_dialog.show()

    def _apply_text_theme(self):
        is_dark = self._current_theme == 'dark'
        bg = self.DARK_TEXT_BG if is_dark else self.LIGHT_TEXT_BG
        fg = self.DARK_TEXT_FG if is_dark else self.LIGHT_TEXT_FG
        for w in (self.app.result_text, self.app.backtest_text):
            w.config(bg=bg, fg=fg, insertbackground=fg)
        if hasattr(self.scanner_ui, 'scanner_result_text'):
            self.scanner_ui.scanner_result_text.config(bg=bg, fg=fg, insertbackground=fg)
        if hasattr(self.scanner_ui, '_legend_text_widget'):
            self.scanner_ui._legend_text_widget.config(bg=bg, fg=fg)
        import matplotlib.pyplot as plt
        plt.style.use('dark_background' if is_dark else 'default')

    def _on_sectors_loaded(self, all_tickers, ticker_to_sector, sector_to_tickers):
        if ticker_to_sector is None:
            self.app.set_tickers_loading(False)
            self.app.result_text.insert(tk.END, "Ошибка загрузки эмитентов (MOEX API недоступен)\n")
            return
        old_count = len(self.sector_db.get_all_tickers())
        self.sector_db.apply_dynamic_data(ticker_to_sector, sector_to_tickers)
        sector_map = self.sector_db.get_ticker_to_sector_map()
        self.app.update_ticker_list(all_tickers, sector_map)
        total = len(self.sector_db.get_tickers(self.sector_db.get_all_sectors()))
        self.scanner_ui.update_total_count(total)
        if hasattr(self, 'smart_scanner_ui'):
            self.smart_scanner_ui._total_tickers = total
            self.smart_scanner_ui._total_count_label.config(text=f"Всего эмитентов: {total}")
            self.smart_scanner_ui.status_var.set(f"Готов к сканированию ({total} эмитентов)")
        added = total - old_count
        self.app.result_text.insert(
            tk.END,
            f"Загружено {total} эмитентов (+{added} из MOEX индексов)\n")

        updated_tickers = self.sector_db.get_all_tickers()
        if hasattr(self, 'watchlist_ui') and self.watchlist_ui is not None and hasattr(self.watchlist_ui, 'ticker_combo'):
            self.watchlist_ui._all_tickers = updated_tickers
            self.watchlist_ui.ticker_combo.configure(values=updated_tickers)
        if hasattr(self, 'pos_calc_ui') and hasattr(self.pos_calc_ui, 'ticker_combo'):
            self.pos_calc_ui._all_tickers = updated_tickers
            self.pos_calc_ui.ticker_combo.configure(values=updated_tickers)
        if hasattr(self, 'signal_journal_ui') and hasattr(self.signal_journal_ui, 'ticker_filter'):
            self.signal_journal_ui._all_tickers = updated_tickers
            self.signal_journal_ui.ticker_filter.configure(values=updated_tickers)

    def _start_auto_load_tickers(self):
        self.app.set_tickers_loading(True)
        self.sector_db.load_dynamic_async(on_complete=self._on_sectors_loaded)

    def _on_toggle_favorite(self, ticker):
        from utils import save_favorites
        self._favorites = toggle_favorite(list(self._favorites), ticker)
        save_favorites(self._favorites)
        return self._favorites

    def _on_close(self) -> None:
        try:
            self.scheduler.stop_all()
        except Exception:
            pass
        try:
            self.alert_monitor.stop()
        except Exception:
            pass
        try:
            self._save_session_state()
        except Exception:
            pass
        try:
            moex_cache.cleanup()
            moex_cache.flush()
        except Exception:
            pass
        try:
            from cloud.sync import sync_manager
            if sync_manager.auto_sync_on_close and sync_manager.is_connected():
                sync_manager.upload_all()
        except Exception:
            pass
        try:
            self.root.destroy()
        except tk.TclError:
            pass

    def get_moex_tickers(self) -> list[str]:
        """Получение списка тикеров MOEX (с кэшем)"""
        return cached_get_tickers(self._fetch_moex_tickers)

    def _fetch_moex_tickers(self) -> list[str]:
        """Получение списка тикеров MOEX (сырой HTTP-запрос)"""
        try:
            url = "https://iss.moex.com/iss/engines/stock/markets/shares/boards/TQBR/securities.json"
            response = _MOEX_SESSION.get(url, timeout=30)
            response.raise_for_status()
            data = response.json()
            if "securities" in data and "data" in data["securities"]:
                return [security[0] for security in data["securities"]["data"]]
            logging.error("Отсутствует securities.data")
            return []
        except Exception as e:
            logging.error(f"Ошибка получения списка акций: {e}")
            return []

    def validate_date(self, date_str: str) -> bool:
        """Проверка формата даты"""
        try:
            datetime.strptime(date_str, "%Y-%m-%d")
            return True
        except ValueError:
            return False

    def export_data_to_txt(self, data: dict, filename: str) -> bool:
        """Экспорт данных в текстовый файл"""
        try:
            with open(filename, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            return True
        except Exception as e:
            logging.error(f"Ошибка экспорта данных: {e}")
            return False

    def _process_stock_data(self, stock_data, start_date, end_date):
        """Обработка загруженных данных (вызывается в главном потоке)"""
        if isinstance(stock_data, str):
            self.app.result_text.delete(1.0, tk.END)
            self.app.result_text.insert(tk.END, f"Ошибка: {stock_data}")
            self.app.get_data_button.config(state='normal', text='1. Получить данные')
            return
        if not isinstance(stock_data, list) or len(stock_data) == 0:
            self.app.result_text.delete(1.0, tk.END)
            self.app.result_text.insert(tk.END, "Ошибка: данные не получены")
            self.app.get_data_button.config(state='normal', text='1. Получить данные')
            return

        self.state.stock_data = stock_data
        self.state.df = self.candles_to_df_custom(stock_data)
        self.app.result_text.delete(1.0, tk.END)

        params = self.app.get_backtest_params()
        min_hits = params.get('min_hits', DEFAULT_MIN_REPEATS) if params else DEFAULT_MIN_REPEATS
        if min_hits < 1:
            min_hits = 1

        self.state.strong_zones = []
        if self.state.df is not None and len(stock_data) > 0:
            atr_series = calc_atr(self.state.df, 14)
            avg_atr = atr_series.mean()
            if not pd.isna(avg_atr) and avg_atr > 0:
                self.state.strong_zones = find_strong_zones(
                    stock_data, atr_value=avg_atr, min_hits=min_hits, max_zones=6)
                for price, count in self.state.strong_zones:
                    self.app.result_text.insert(
                        tk.END, f"Зона: {price:.2f}, касаний: {count}\n")
            else:
                self.app.result_text.insert(tk.END, "Недостаточно данных для ATR\n")

        self.state.current_step = 0
        self.app.get_data_button.config(state='normal', text='1. Получить данные')

        end_dt = datetime.strptime(end_date, "%Y-%m-%d")
        start_dt = datetime.strptime(start_date, "%Y-%m-%d")
        total_days = (end_dt - start_dt).days
        self.app.result_text.insert(tk.END, f"\nСвечей: {len(stock_data)}, период: ~{total_days} дн.")
        self._save_session_state()
        self.app.update_chart(self.state.df, strong_zones=self.state.strong_zones)

    def on_select(self) -> None:
        """Обработка выбора акции и дат (асинхронная загрузка)"""
        selected_stock = self.app.get_selected_ticker()
        start_date = self.app.start_date_entry.get()
        end_date = self.app.end_date_entry.get()

        if not self.validate_date(start_date):
            self.app.result_text.delete(1.0, tk.END)
            self.app.result_text.insert(tk.END, "Ошибка: неверный формат начальной даты")
            return
        if not self.validate_date(end_date):
            self.app.result_text.delete(1.0, tk.END)
            self.app.result_text.insert(tk.END, "Ошибка: неверный формат конечной даты")
            return

        self.app.get_data_button.config(state='disabled', text='Загрузка...')
        self.app.result_text.delete(1.0, tk.END)
        self.app.result_text.insert(tk.END, "Загрузка данных с MOEX...")

        def fetch_task():
            try:
                data = self.get_stock_data(selected_stock, start_date, end_date)
                self.root.after(0, lambda: self._process_stock_data(
                    data, start_date, end_date))
            except Exception as e:
                self.root.after(0, lambda: self._on_select_error(str(e)))

        t = threading.Thread(target=fetch_task, daemon=True)
        t.start()

    def _on_select_error(self, error_msg):
        self.app.result_text.delete(1.0, tk.END)
        self.app.result_text.insert(tk.END, f"Ошибка: {error_msg}")
        self.app.get_data_button.config(state='normal', text='1. Получить данные')

    def on_export_button(self) -> None:
        """Экспорт данных в текстовый файл"""
        if self.state.stock_data is None:
            print("Ошибка: данные по акции не загружены")
            return
        selected_stock = self.app.get_selected_ticker()
        if isinstance(self.state.stock_data, list):
            filename = f"{selected_stock}_data.txt"
            if self.export_data_to_txt(self.state.stock_data, filename):
                print(f"Data exported to {filename}")
            else:
                print("Failed to export data.")

    def _run_backtest_task(self, params):
        try:
            engine = BacktestEngine(**params)
            selected_stock = self.app.get_selected_ticker()
            strategy_name = params.get('strategy', '')

            def _record_signal(ticker, side, price, sl, tp, entered, date=None):
                try:
                    if date is not None and hasattr(date, 'strftime'):
                        date_str = date.strftime('%Y-%m-%d %H:%M')
                    elif date is not None:
                        date_str = str(date)
                    else:
                        date_str = None
                    self.signal_storage.add_signal(
                        ticker=selected_stock or ticker,
                        side=side,
                        price=price,
                        strategy=strategy_name,
                        sl=sl,
                        tp=tp,
                        entered=entered,
                        date=date_str,
                    )
                except Exception:
                    pass

            engine.signal_recorder = _record_signal
            trades, metrics = engine.run(self.state.stock_data)
            selected_stock = self.app.get_selected_ticker()
            self._last_trades = trades
            self._last_metrics = metrics
            self._last_export_stock = selected_stock

            signal = None
            stock_data = self.state.stock_data
            if trades and stock_data:
                last_price = float(stock_data[-1][1])
                df = candles_to_df(stock_data)
                if df is not None and len(df) > 0:
                    atr_series = calc_atr(df, params.get('atr_period', 14))
                    if not atr_series.empty and not atr_series.isna().iloc[-1]:
                        atr_value = atr_series.iloc[-1]
                        last_candles = DEFAULT_LAST_CANDLES
                        levels_strength = calculate_level_strength(trades, last_candles=last_candles)
                        if levels_strength and last_price and atr_value:
                            atr_sl_val = params.get('atr_sl', 1.0)
                            atr_tp_val = params.get('atr_tp', 2.0)
                            signal = get_best_level_signal(
                                levels_strength, last_price, atr_value,
                                atr_sl=atr_sl_val, atr_tp=atr_tp_val)
                            if signal and signal['action'] == 'NONE':
                                signal = get_best_level_signal(
                                    levels_strength, last_price, atr_value,
                                    threshold_mult=1.0, atr_sl=atr_sl_val, atr_tp=atr_tp_val)
                            if signal:
                                signal['last_price'] = last_price
                                signal['atr'] = atr_value

            self.root.after(0, lambda e=engine: self._on_backtest_complete(
                trades, metrics, selected_stock, signal, params, engine=e))
        except Exception as e:
            self.root.after(0, lambda: self._on_backtest_error(str(e)))

    def _on_backtest_complete(self, trades, metrics,
                              selected_stock, signal=None, params=None, engine=None):
        self.app.backtest_button.config(state='normal', text='2. Backtest')
        self.app.display_backtest_results(metrics, params)
        self.app.enable_save_results_button()
        self._last_trades = trades
        if params and 'capital' in params:
            self._last_capital = params['capital']
        if signal:
            self.app.display_recommendation(signal, params)
            self.app.set_last_analysis(signal, params)
        engine_levels = engine.last_levels if engine else []
        if signal:
            sl_price = signal.get('sl_price')
            tp_price = signal.get('tp_price')
        else:
            sl_price = tp_price = None
        self.app.add_post_backtest_lines(
            engine_levels=engine_levels,
            sl_price=sl_price,
            tp_price=tp_price)

        if metrics and metrics.get('max_drawdown', 0) > 0:
            self.notification_manager.check_drawdown(metrics['max_drawdown'])

    def _on_backtest_error(self, error_msg):
        self.app.add_backtest_result(f"Ошибка backtest: {error_msg}")
        self.app.backtest_button.config(state='normal', text='2. Backtest')
        logging.exception("Backtest error")

    def _on_save_results(self):
        if not self._last_trades and not self._last_metrics:
            return
        stock = self._last_export_stock or self.app.get_selected_ticker()
        csv_path, json_path = _export_results(
            self._last_trades or [], self._last_metrics or {}, stock)
        self.app.backtest_text.insert(tk.END, f"\n\nФайлы:\n  {csv_path}\n  {json_path}")

    def on_backtest(self) -> None:
        """Запуск backtesting стратегии (асинхронно)"""
        if self.state.stock_data is None or not isinstance(self.state.stock_data, list):
            self.app.add_backtest_result("Ошибка: сначала загрузите данные (кнопка 1)")
            return
        if len(self.state.stock_data) < MIN_CANDLES_FOR_BACKTEST:
            self.app.add_backtest_result("Ошибка: слишком мало данных (нужно >= 30 свечей)")
            return

        params = self.app.get_backtest_params()
        if params is None:
            self.app.add_backtest_result("Ошибка: проверьте числовые параметры.")
            return

        self.app.backtest_button.config(state='disabled', text='Backtest...')

        t = threading.Thread(
            target=self._run_backtest_task,
            args=(params,),
            daemon=True
        )
        t.start()

    def on_optimize(self) -> None:
        """Запуск оптимизации параметров стратегии (асинхронно)"""
        if self.state.stock_data is None or not isinstance(self.state.stock_data, list):
            self.app.add_backtest_result("Ошибка: сначала загрузите данные (кнопка 1)")
            return
        if len(self.state.stock_data) < MIN_CANDLES_FOR_BACKTEST:
            self.app.add_backtest_result("Ошибка: слишком мало данных (нужно >= 30 свечей)")
            return

        params = self.app.get_backtest_params()
        if params is None:
            self.app.add_backtest_result("Ошибка: проверьте числовые параметры.")
            return

        strategy_id = params.pop('strategy', 'bounce')

        self.app.optimize_button.config(state='disabled', text='Оптимизация...')

        def run():
            try:
                from optimization.grid import optimize
                results, total = optimize(
                    strategy_id,
                    self.state.stock_data,
                    default_params=params,
                    progress_fn=lambda c, t: None
                )
                self.root.after(0, lambda: self._on_optimize_complete(results, total, strategy_id))
            except Exception as e:
                self.root.after(0, lambda: self._on_optimize_error(str(e)))

        t = threading.Thread(target=run, daemon=True)
        t.start()

    def _on_optimize_complete(self, results, total, strategy_id):
        self.app.optimize_button.config(state='normal', text='3. Оптимизация')
        from strategy.config import STRATEGY_REGISTRY
        strat_name = STRATEGY_REGISTRY.get(strategy_id, {}).get('name', strategy_id)

        txt = self.app.backtest_text
        txt.delete(1.0, tk.END)
        txt.insert(tk.END, (
            f"========== ОПТИМИЗАЦИЯ: {strat_name} ==========\n"
            f"Проверено комбинаций: {total}\n"
            f"Лучших результатов: {len(results)}\n\n"
        ))

        if results:
            best = results[0]
            txt.insert(tk.END, "── Лучшие параметры ──\n")
            for k, v in best['params'].items():
                txt.insert(tk.END, f"  {k}: {v}\n")
            txt.insert(tk.END, "\n")
            txt.insert(tk.END, (
                f"  Sharpe:       {best['sharpe']:.2f}\n"
                f"  Profit Factor:{best['profit_factor']:.2f}\n"
                f"  Доходность:   {best['total_return']:+.2f}%\n"
                f"  Max Drawdown: -{best['max_drawdown']:.2f}%\n"
                f"  Сделок:       {best['total_trades']}\n"
                f"  Win Rate:     {best['win_rate']:.1f}%\n"
            ))
            if 'oos_sharpe' in best:
                degrad = best.get('degradation', 1) * 100
                txt.insert(tk.END, (
                    f"\n  ── OOS (Out-Of-Sample) ──\n"
                    f"  OOS Sharpe:   {best['oos_sharpe']:.2f}\n"
                    f"  OOS Return:   {best['oos_return']:+.2f}%\n"
                    f"  OOS Drawdown: -{best['oos_drawdown']:.2f}%\n"
                    f"  OOS Сделок:   {best['oos_trades']}\n"
                    f"  Деградация:   {degrad:.0f}%\n"
                ))
            txt.insert(tk.END, "\n")
            txt.insert(tk.END, "── Топ-5 комбинаций (нажмите [Применить]) ──\n")
            for rank, r in enumerate(results[:5], 1):
                param_str = ", ".join(f"{k}={v}" for k, v in r['params'].items())
                txt.insert(tk.END, f"  {rank}. ")
                tag_name = f"apply_{rank - 1}"
                txt.insert(tk.END, "[Применить] ", (tag_name,))
                oos_info = ""
                if 'oos_sharpe' in r:
                    oos_info = f" OOS:{r['oos_sharpe']:.2f}/{r['oos_return']:+.0f}% D:{r.get('degradation', 1)*100:.0f}%"
                txt.insert(tk.END, (
                    f"Sharpe={r['sharpe']:.2f} PF={r['profit_factor']:.2f} "
                    f"Ret={r['total_return']:+.1f}%{oos_info} | {param_str}\n"
                ))
                txt.tag_config(tag_name, foreground='blue', underline=1)
                txt.tag_bind(tag_name, '<Button-1>',
                             lambda e, idx=rank - 1: self._apply_optimized_params(results[idx]['params']))
                txt.tag_bind(tag_name, '<Enter>',
                             lambda e, tn=tag_name: txt.tag_config(tn, foreground='#0066ff'))
                txt.tag_bind(tag_name, '<Leave>',
                             lambda e, tn=tag_name: txt.tag_config(tn, foreground='blue'))
        else:
            txt.insert(tk.END, "Не найдено комбинаций с достаточным числом сделок (>=5).\n")
        txt.insert(tk.END, "=" * 50)

    def _apply_optimized_params(self, params_dict):
        """Применить параметры из оптимизации к текущим настройкам бумаги."""
        app = self.app
        txt = app.backtest_text

        try:
            app._rebuild_params()
            for key, value in params_dict.items():
                entry = app._param_entries.get(key)
                if entry is None:
                    continue
                if isinstance(entry, ttk.Combobox):
                    try:
                        entry.current(int(value))
                    except (ValueError, TypeError):
                        pass
                else:
                    entry.delete(0, tk.END)
                    entry.insert(0, str(value))

            app._save_current_settings()
            self.root.update_idletasks()

            ticker = app._extract_ticker(app.stock_combobox.get())
            txt.delete(1.0, tk.END)
            txt.insert(tk.END, (
                f"Параметры применены к бумаге {ticker}.\n"
                "Настройки сохранены в ticker_settings.json.\n\n"
                "Применённые параметры:\n"
            ))
            for k, v in params_dict.items():
                txt.insert(tk.END, f"  {k}: {v}\n")
        except Exception as e:
            txt.delete(1.0, tk.END)
            txt.insert(tk.END, f"Ошибка при применении параметров: {e}")
            raise

    def _on_optimize_error(self, error_msg):
        self.app.add_backtest_result(f"Ошибка оптимизации: {error_msg}")
        self.app.optimize_button.config(state='normal', text='3. Оптимизация')

    def on_walkforward(self) -> None:
        """Запуск Walk-forward анализа."""
        if self.state.stock_data is None or not isinstance(self.state.stock_data, list):
            self.app.add_backtest_result("Ошибка: сначала загрузите данные (кнопка 1)")
            return
        if len(self.state.stock_data) < MIN_CANDLES_FOR_BACKTEST:
            self.app.add_backtest_result("Ошибка: слишком мало данных (нужно >= 30 свечей)")
            return

        params = self.app.get_backtest_params()
        if params is None:
            self.app.add_backtest_result("Ошибка: проверьте числовые параметры.")
            return

        strategy_id = params.pop('strategy', 'bounce')

        self.app.walkforward_button.config(state='disabled', text='Walk-fwd...')
        txt = self.app.backtest_text
        txt.delete(1.0, tk.END)
        txt.insert(tk.END, "Запуск Walk-forward анализа...\n")
        self.root.update_idletasks()

        def run():
            try:
                from optimization.walkforward import run_walkforward, summarize_walkforward
                results = run_walkforward(
                    strategy_id,
                    self.state.stock_data,
                    window_years=2,
                    step_years=1,
                    oos_split=0.3,
                    default_params=params,
                    progress_fn=lambda c, t: None,
                )
                report = summarize_walkforward(results)
                self.root.after(0, lambda: self._on_walkforward_complete(report))
            except Exception as e:
                self.root.after(0, lambda: self._on_walkforward_error(str(e)))

        t = threading.Thread(target=run, daemon=True)
        t.start()

    def _on_walkforward_complete(self, report):
        self.app.walkforward_button.config(state='normal', text='5. Walk-fwd')
        txt = self.app.backtest_text
        txt.delete(1.0, tk.END)
        txt.insert(tk.END, report)

    def _on_walkforward_error(self, error_msg):
        self.app.add_backtest_result(f"Ошибка Walk-forward: {error_msg}")
        self.app.walkforward_button.config(state='normal', text='5. Walk-fwd')

    def on_sensitivity(self) -> None:
        if self.state.stock_data is None or not isinstance(self.state.stock_data, list):
            self.app.add_backtest_result("Ошибка: сначала загрузите данные (кнопка 1)")
            return
        if len(self.state.stock_data) < MIN_CANDLES_FOR_BACKTEST:
            self.app.add_backtest_result("Ошибка: слишком мало данных (нужно >= 30 свечей)")
            return

        params = self.app.get_backtest_params()
        if params is None:
            return

        strategy_id = params.get('strategy', 'breakout')
        self.app.sensitivity_button.config(state='disabled', text='Анализ...')
        self.app.add_backtest_result("Анализ чувствительности... подождите")

        def run():
            from optimization.sensitivity import analyze_sensitivity, format_sensitivity_report, plot_tornado
            try:
                result = analyze_sensitivity(
                    strategy_id=strategy_id,
                    candles=self.state.stock_data,
                    base_params=params,
                )
                report = format_sensitivity_report(result)
                self.root.after(0, lambda: self._on_sensitivity_complete(report, result))
            except Exception as e:
                self.root.after(0, lambda: self._on_sensitivity_error(str(e)))

        import threading
        threading.Thread(target=run, daemon=True).start()

    def _on_sensitivity_complete(self, report, result):
        self.app.sensitivity_button.config(state='normal', text='6. Чувств.')
        txt = self.app.backtest_text
        txt.delete(1.0, tk.END)
        txt.insert(tk.END, report)

        try:
            from optimization.sensitivity import plot_tornado
            fig = plot_tornado(result)
            if fig is not None:
                try:
                    import matplotlib
                    matplotlib.use('TkAgg')
                    fig.show()
                except Exception:
                    pass
        except Exception:
            pass

    def _on_sensitivity_error(self, error_msg):
        self.app.add_backtest_result(f"Ошибка анализа чувствительности: {error_msg}")
        self.app.sensitivity_button.config(state='normal', text='6. Чувств.')

    def on_heatmap(self) -> None:
        """2D Sensitivity Heatmap: открывает окно выбора двух параметров."""
        if self.state.stock_data is None or not isinstance(self.state.stock_data, list):
            self.app.add_backtest_result("Ошибка: сначала загрузите данные (кнопка 1)")
            return
        if len(self.state.stock_data) < MIN_CANDLES_FOR_BACKTEST:
            self.app.add_backtest_result("Ошибка: слишком мало данных (нужно >= 30 свечей)")
            return

        try:
            import tkinter as tk
            from tkinter import ttk
        except ImportError:
            self.app.add_backtest_result("Ошибка: tkinter недоступен")
            return

        params = self.app.get_backtest_params()
        if params is None:
            self.app.add_backtest_result("Ошибка: проверьте числовые параметры.")
            return

        strategy_id = params.pop('strategy', 'bounce')

        # Build param list from strategy config
        from strategy.config import get_strategy_params
        sp = get_strategy_params(strategy_id)
        tunable = [p['key'] for p in sp
                   if p['key'] not in ('strategy', 'capital', 'risk_per_trade',
                                        'commission', 'entry_type', 'use_pivot_levels',
                                        'use_mtf_filter', 'position_sizing',
                                        'trailing_sl', 'partial_tp')
                   and p.get('type') in (int, float)]
        # keep only those that are in params and non-zero
        tunable = [k for k in tunable if k in params and params.get(k, 0) != 0]
        if not tunable:
            self.app.add_backtest_result("Нет варьируемых параметров.")
            return

        # Dialog window
        win = tk.Toplevel(self.root)
        win.title('Heatmap параметров')
        win.geometry('400x300')
        win.transient(self.root)
        win.grab_set()

        ttk.Label(win, text='Параметр X:').pack(pady=(15, 2))
        var_x = tk.StringVar(value=tunable[0])
        cx = ttk.Combobox(win, textvariable=var_x, values=tunable, state='readonly')
        cx.pack(fill='x', padx=20)

        ttk.Label(win, text='Параметр Y:').pack(pady=(10, 2))
        var_y = tk.StringVar(value=tunable[1] if len(tunable) > 1 else tunable[0])
        cy = ttk.Combobox(win, textvariable=var_y, values=tunable, state='readonly')
        cy.pack(fill='x', padx=20)

        ttk.Label(win, text='Метрика:').pack(pady=(10, 2))
        var_m = tk.StringVar(value='sharpe')
        cm = ttk.Combobox(win, textvariable=var_m,
                          values=['sharpe', 'total_return', 'profit_factor',
                                  'max_drawdown', 'win_rate'],
                          state='readonly')
        cm.pack(fill='x', padx=20)

        ttk.Label(win, text='Шагов сетки:').pack(pady=(10, 2))
        var_n = tk.IntVar(value=6)
        cn = ttk.Combobox(win, textvariable=var_n,
                          values=[4, 5, 6, 8, 10, 12], state='readonly')
        cn.pack(fill='x', padx=20)

        def run_heatmap():
            px = var_x.get()
            py = var_y.get()
            if px == py:
                self.app.add_backtest_result("Параметры X и Y должны различаться.")
                return
            metric = var_m.get()
            n_steps = var_n.get()
            win.destroy()
            self._do_heatmap(strategy_id, params, px, py, metric, n_steps)

        ttk.Button(win, text='Построить Heatmap', command=run_heatmap).pack(pady=20)

    def _do_heatmap(self, strategy_id, params, param_x, param_y, metric, n_steps):
        from optimization.sensitivity import plot_heatmap_grid
        try:
            fig = plot_heatmap_grid(
                strategy_id, self.state.stock_data, param_x, param_y,
                base_params=params, n_steps=n_steps, metric=metric)
            if fig is not None:
                fig.show()
            else:
                self.app.add_backtest_result("Heatmap не построен (недостаточно данных).")
        except Exception as e:
            self.app.add_backtest_result(f"Ошибка построения Heatmap: {e}")

    def _on_fundamental(self):
        ticker = self.app.get_selected_ticker()
        if not ticker:
            self.app.add_backtest_result("Ошибка: выберите тикер")
            return

        self.app.fundamental_button.config(state='disabled', text='Загрузка...')
        self.app.add_backtest_result(f"Загрузка фундаментальных данных {ticker}...")

        def run():
            from fundamental.data import fetch_all_tqbr_fundamentals, fetch_dividends, calc_dividend_yield, format_fundamental_report
            from fundamental.scoring import compute_total_score, format_score_report
            try:
                all_data = fetch_all_tqbr_fundamentals()
                td = all_data.get(ticker, {})
                price = td.get('prev_price')
                market_cap = td.get('market_cap')
                shares = td.get('issue_size')

                dividends = fetch_dividends(ticker)
                div_yield = calc_dividend_yield(dividends, price) if dividends and price else None

                fund_data = {
                    'ticker': ticker,
                    'price': price,
                    'market_cap': market_cap,
                    'shares': shares,
                    'div_yield': div_yield,
                    'dividends': dividends,
                    'last_div': dividends[0] if dividends else None,
                }

                report = format_fundamental_report(fund_data)

                total, details, rating = compute_total_score(fund_data)
                score_report = format_score_report(total, details, rating)

                full_report = report + '\n\n' + score_report
                self.root.after(0, lambda: self._on_fundamental_complete(full_report))
            except Exception as e:
                self.root.after(0, lambda: self._on_fundamental_error(str(e)))

        import threading
        threading.Thread(target=run, daemon=True).start()

    def _on_fundamental_complete(self, report):
        self.app.fundamental_button.config(state='normal', text='7. Фундам.')
        txt = self.app.backtest_text
        txt.delete(1.0, tk.END)
        txt.insert(tk.END, report)

    def _on_fundamental_error(self, error_msg):
        self.app.add_backtest_result(f"Ошибка фундаментального анализа: {error_msg}")
        self.app.fundamental_button.config(state='normal', text='7. Фундам.')

    def on_portfolio(self) -> None:
        """Портфельный бэктест по тикерам из торгового дневника."""
        entries = self.diary_storage.load()
        if not entries:
            self.app.add_backtest_result("Ошибка: дневник пуст. Добавьте сделки в дневник.")
            return

        tickers = sorted({e.ticker for e in entries})
        if len(tickers) < 2:
            self.app.add_backtest_result("Ошибка: в дневнике меньше 2 уникальных тикеров.")
            return

        params = self.app.get_backtest_params()
        if params is None:
            self.app.add_backtest_result("Ошибка: проверьте числовые параметры.")
            return
        strategy_id = params.pop('strategy', 'bounce')

        risk_params = self.app.get_portfolio_risk_params()
        if risk_params is None:
            self.app.add_backtest_result("Ошибка: проверьте параметры лимитов портфеля.")
            return

        start_date = self.app.start_date_entry.get()
        end_date = self.app.end_date_entry.get()

        self.app.portfolio_button.config(state='disabled', text='Портфель...')
        txt = self.app.backtest_text
        txt.delete(1.0, tk.END)
        txt.insert(tk.END, f"Загрузка данных для {len(tickers)} тикеров из дневника: {', '.join(tickers)}\n")
        self.root.update_idletasks()

        sector_map = self.sector_db.get_ticker_to_sector_map()

        def fetch_and_run():
            try:
                portfolio_data = {}
                errors = []
                for t in tickers:
                    data = self.get_stock_data(t, start_date, end_date)
                    if isinstance(data, str):
                        errors.append(f"{t}: {data}")
                    elif len(data) < MIN_CANDLES_FOR_BACKTEST:
                        errors.append(f"{t}: недостаточно данных ({len(data)} свечей)")
                    else:
                        portfolio_data[t] = data

                if not portfolio_data:
                    self.root.after(0, lambda: self._on_portfolio_error("Нет данных ни по одному тикеру."))
                    return

                from backtest.portfolio import run_portfolio
                from backtest.portfolio_risk import PortfolioRiskManager

                risk_manager = PortfolioRiskManager(
                    max_open_positions=risk_params['max_open_positions'],
                    max_drawdown_pct=risk_params['max_drawdown_pct'],
                    cooldown_bars=risk_params['cooldown_bars'],
                    max_sector_exposure=risk_params['max_sector_exposure'],
                    sector_map=sector_map,
                )

                result = run_portfolio(portfolio_data, capital=params.get('capital', 1_000_000),
                                       risk_manager=risk_manager, **params)

                self.root.after(0, lambda: self._on_portfolio_complete(result, strategy_id, errors))
            except Exception as e:
                self.root.after(0, lambda: self._on_portfolio_error(str(e)))

        t = threading.Thread(target=fetch_and_run, daemon=True)
        t.start()

    def _on_portfolio_complete(self, result, strategy_id, errors):
        self.app.portfolio_button.config(state='normal', text='4. Портфель')
        pm = result['portfolio_metrics']
        txt = self.app.backtest_text
        txt.delete(1.0, tk.END)

        from strategy.config import STRATEGY_REGISTRY
        strat_name = STRATEGY_REGISTRY.get(strategy_id, {}).get('name', strategy_id)

        lines = [
            f"========== ПОРТФЕЛЬНЫЙ БЭКТЕСТ ==========",
            f"Стратегия: {strat_name}",
            f"Тикеров: {len(result['ticker_results'])}",
            "",
            f"Начальный капитал: {pm['initial_capital']:,.0f} руб",
            f"Конечный капитал:   {pm['final_capital']:,.0f} руб",
            f"Чистая прибыль:     {pm['net_profit']:+,.0f} руб",
            f"Общая доходность:   {pm['total_return']:+.2f} %",
            "",
            f"Всего сделок:       {pm['total_trades']}",
            f"Win Rate:           {pm['win_rate']:.1f} %",
            f"Profit Factor:      {pm['profit_factor']}",
            f"Max Drawdown:       -{pm['max_drawdown']:.2f} %",
            f"Sharpe Ratio:       {pm['sharpe']}",
        ]

        risk_stats = result.get('risk_stats')
        if risk_stats:
            lines.append("")
            lines.append("── Лимиты портфеля ──")
            if risk_stats.get('blocked_positions', 0) > 0:
                lines.append(f"  Заблокировано по лимиту позиций: {risk_stats['blocked_positions']}")
            if risk_stats.get('blocked_drawdown', 0) > 0:
                lines.append(f"  Заблокировано по стоп-просадке: {risk_stats['blocked_drawdown']}")
            if risk_stats.get('blocked_sector', 0) > 0:
                lines.append(f"  Заблокировано по сектору: {risk_stats['blocked_sector']}")
            if risk_stats.get('drawdown_halts', 0) > 0:
                lines.append(f"  Срабатываний портфельного стопа: {risk_stats['drawdown_halts']}")
            if not any(v > 0 for k, v in risk_stats.items() if k != 'drawdown_halts'):
                lines.append("  Лимиты не сработали (нет блокировок)")

        lines.append("")
        lines.append("── По тикерам ──")
        for ticker, tr in sorted(result['ticker_results'].items()):
            m = tr['metrics']
            lines.append(
                f"  {ticker}: {m.get('total_return', 0):+.2f}% | "
                f"сделок={m.get('total_trades', 0)} | "
                f"Sharpe={m.get('sharpe', 0)} | "
                f"DD=-{m.get('max_drawdown', 0):.1f}%"
            )

        if errors:
            lines.append("")
            lines.append("── Ошибки ──")
            for e in errors:
                lines.append(f"  {e}")

        lines.append("=" * 50)
        txt.insert(tk.END, '\n'.join(lines))

    def _on_portfolio_error(self, error_msg):
        self.app.add_backtest_result(f"Ошибка портфельного бэктеста: {error_msg}")
        self.app.portfolio_button.config(state='normal', text='4. Портфель')

    def on_scanner(self) -> None:
        """Запуск сканера по секторам"""
        selected = self.scanner_ui.get_selected_sectors()
        if not selected:
            self.scanner_ui.show_report("Ошибка: выберите хотя бы один сектор.")
            return

        params = self.scanner_ui.get_backtest_params()
        if params is None:
            self.scanner_ui.show_report("Ошибка: проверьте числовые параметры.")
            return

        date_from = self.scanner_ui.scanner_date_from.get()
        date_to = self.scanner_ui.scanner_date_to.get()
        if not self.validate_date(date_from) or not self.validate_date(date_to):
            self.scanner_ui.show_report("Ошибка: проверьте формат дат (гггг-мм-дд).")
            return

        fund_filter = self.scanner_ui.get_fund_filter()
        self._last_scan_params = params

        def run_scan():
            try:
                scanner = Scanner(sector_db=self.sector_db, fetch_fn=self.get_stock_data)
                ticker_settings_path = os.path.join(app_dir(), 'results', 'ticker_settings.json')
                results = scanner.scan(
                    sectors=selected,
                    date_from=date_from,
                    date_to=date_to,
                    backtest_params=params,
                    ticker_settings_path=ticker_settings_path,
                    progress_fn=lambda c, t, tick, sec: self.scanner_ui.update_progress(c, t, tick, sec)
                )

                if fund_filter:
                    from fundamental.data import fetch_all_tqbr_fundamentals, fetch_dividends, calc_dividend_yield
                    from fundamental.filters import filter_by_fundamentals
                    result_tickers = {r.get('ticker') for r in results if r.get('ticker')}
                    all_fund = fetch_all_tqbr_fundamentals()
                    for ticker in result_tickers:
                        td = all_fund.get(ticker)
                        if not td:
                            continue
                        price = td.get('prev_price')
                        if price:
                            try:
                                divs = fetch_dividends(ticker)
                                td['div_yield'] = calc_dividend_yield(divs, price) if divs else None
                                td['dividends'] = divs
                                td['div_count'] = len(divs)
                            except Exception:
                                pass
                    filtered_tickers = filter_by_fundamentals(all_fund, fund_filter)
                    results = [r for r in results if r.get('ticker') in filtered_tickers]

                self._last_scan_results = results
                n_custom = len(scanner.ticker_overrides_used)
                report = generate_report(results, top_n=5, params=params)
                if n_custom:
                    report += f"\n  Бумаг с индивид. настройками: {n_custom}"
                if fund_filter:
                    report += f"\n  Фундаментальный фильтр применён"
                self.root.after(0, lambda: self.scanner_ui.show_report(report))
            except Exception as e:
                self.root.after(0, lambda: self.scanner_ui.show_report(f"Ошибка: {str(e)}"))
            finally:
                self.root.after(0, lambda: self.scanner_ui.set_running(False))

        self.scanner_ui.set_running(True)
        t = threading.Thread(target=run_scan, daemon=True)
        t.start()

    def on_export_excel(self) -> None:
        """Экспорт результатов сканера в Excel"""
        if not self._last_scan_results:
            self.scanner_ui.show_report("Ошибка: сначала запустите сканер.")
            return
        from screening.reporter import export_to_excel
        os.makedirs(os.path.join(app_dir(), 'results'), exist_ok=True)
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        fpath = os.path.join(app_dir(), f'results/scanner_{ts}.xlsx')
        try:
            export_to_excel(self._last_scan_results, self._last_scan_params, fpath)
            self.scanner_ui.scanner_result_text.insert(tk.END, f"\n\nExcel сохранён: {fpath}")
        except ImportError:
            self.scanner_ui.show_report("Ошибка: установите openpyxl (pip install openpyxl)")
        except Exception as e:
            self.scanner_ui.show_report(f"Ошибка экспорта: {str(e)}")

    def on_smart_scanner(self) -> None:
        """Запуск умного сканера (все стратегии на каждом тикере)"""
        selected = self.smart_scanner_ui.get_selected_sectors()
        if not selected:
            self.smart_scanner_ui.show_results([])
            return

        params = self.smart_scanner_ui.get_backtest_params()
        if params is None:
            self.smart_scanner_ui.show_results([])
            return

        date_from = self.smart_scanner_ui.smart_date_from.get()
        date_to = self.smart_scanner_ui.smart_date_to.get()
        if not self.validate_date(date_from) or not self.validate_date(date_to):
            self.smart_scanner_ui.show_results([])
            return

        min_trades = params.pop('min_trades', 30)
        self._last_smart_params = params

        def run_scan():
            try:
                scanner = SmartScanner(sector_db=self.sector_db, fetch_fn=self.get_stock_data)
                results = scanner.scan(
                    sectors=selected,
                    date_from=date_from,
                    date_to=date_to,
                    base_params=params,
                    min_trades=min_trades,
                    progress_fn=lambda c, t, tick, sid_name: self.smart_scanner_ui.update_progress(c, t, tick, sid_name)
                )
                self._last_smart_results = results
                skipped = getattr(scanner, 'skipped_count', 0)
                tested = len(results)
                if skipped > 0 and tested == 0:
                    msg = f"Сканирование завершено. Все тикеры пропущены ({skipped}). Проверьте подключение к MOEX."
                elif skipped > 0:
                    msg = f"Готово. Найдено: {tested}, пропущено: {skipped}"
                else:
                    msg = None
                self.root.after(0, lambda: self.smart_scanner_ui.show_results(results))
                if msg:
                    self.root.after(0, lambda: self.smart_scanner_ui.status_var.set(msg))
            except Exception as e:
                self.root.after(0, lambda: self.smart_scanner_ui.status_var.set(f"Ошибка: {str(e)}"))
            finally:
                self.root.after(0, lambda: self.smart_scanner_ui.set_running(False))

        self.smart_scanner_ui.set_running(True)
        t = threading.Thread(target=run_scan, daemon=True)
        t.start()

    def on_smart_export_excel(self) -> None:
        """Экспорт результатов умного сканера в Excel"""
        if not self._last_smart_results:
            self.smart_scanner_ui.status_var.set("Ошибка: сначала запустите умный сканер.")
            return
        from screening.reporter import export_smart_scan_excel
        os.makedirs(os.path.join(app_dir(), 'results'), exist_ok=True)
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        fpath = os.path.join(app_dir(), f'results/smart_scanner_{ts}.xlsx')
        try:
            export_smart_scan_excel(self._last_smart_results, self._last_smart_params, fpath)
            self.smart_scanner_ui.status_var.set(f"Excel сохранён: {fpath}")
        except ImportError:
            self.smart_scanner_ui.status_var.set("Ошибка: установите openpyxl (pip install openpyxl)")
        except Exception as e:
            self.smart_scanner_ui.status_var.set(f"Ошибка экспорта: {str(e)}")

    def on_add_to_diary(self) -> None:
        """Добавить выбранные сигналы в торговый дневник"""
        results = self._last_scan_results
        if not results:
            self.scanner_ui.show_report("Ошибка: сначала запустите сканер.")
            return

        actionable = [r for r in results if r['signal']['action'] in ('BUY', 'SELL')]
        if not actionable:
            self.scanner_ui.show_report("Нет сигналов BUY/SELL для добавления.")
            return

        params = self._last_scan_params
        capital = params.get('capital', 1_000_000)
        risk_per_trade = params.get('risk_per_trade', 0.02)
        max_hold = params.get('max_hold', 20)

        win = tk.Toplevel(self.root)
        win.title("Добавить в дневник")
        win.geometry("750x500")
        win.transient(self.root)
        win.grab_set()

        frame = ttk.Frame(win)
        frame.pack(fill=tk.BOTH, expand=1, padx=10, pady=10)

        vars_map = {}

        list_frame = ttk.Frame(frame)
        list_frame.pack(fill=tk.BOTH, expand=1)

        canvas = tk.Canvas(list_frame)
        scrollbar = ttk.Scrollbar(list_frame, orient='vertical', command=canvas.yview)
        scrollable = ttk.Frame(canvas)

        scrollable.bind('<Configure>', lambda e: canvas.configure(scrollregion=canvas.bbox('all')))
        canvas.create_window((0, 0), window=scrollable, anchor='nw')
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=1)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        ttk.Label(scrollable, text='',
                  font=('', 9, 'bold')).grid(row=0, column=0, sticky='w', padx=2)
        ttk.Label(scrollable, text='Тикер', font=('', 9, 'bold')).grid(
            row=0, column=1, sticky='w', padx=5)
        ttk.Label(scrollable, text='Сигнал', font=('', 9, 'bold')).grid(
            row=0, column=2, sticky='w', padx=5)
        ttk.Label(scrollable, text='Цена', font=('', 9, 'bold')).grid(
            row=0, column=3, sticky='w', padx=5)
        ttk.Label(scrollable, text='SL', font=('', 9, 'bold')).grid(
            row=0, column=4, sticky='w', padx=5)
        ttk.Label(scrollable, text='TP', font=('', 9, 'bold')).grid(
            row=0, column=5, sticky='w', padx=5)
        ttk.Label(scrollable, text='Объём (₽)', font=('', 9, 'bold')).grid(
            row=0, column=6, sticky='w', padx=5)

        for idx, r in enumerate(actionable, 1):
            sig = r['signal']
            ticker = r['ticker']
            action = sig['action']
            entry_price = r.get('last_price') or sig['level'] or 0
            sl_price = sig.get('sl_price', 0)
            tp_price = sig.get('tp_price', 0)
            qty = calc_position_qty(capital, risk_per_trade, entry_price, sl_price)
            volume = calc_position_volume(capital, risk_per_trade, entry_price, sl_price)

            var = tk.BooleanVar(value=False)
            vars_map[ticker] = {
                'var': var,
                'ticker': ticker,
                'action': action,
                'entry_price': entry_price,
                'sl_price': sl_price,
                'tp_price': tp_price,
                'qty': qty,
                'volume': volume,
                'date': datetime.now().strftime('%Y-%m-%d %H:%M')
            }

            cb = ttk.Checkbutton(scrollable, variable=var)
            cb.grid(row=idx, column=0, padx=2, pady=1)

            ttk.Label(scrollable, text=ticker).grid(
                row=idx, column=1, sticky='w', padx=5)
            ttk.Label(scrollable, text=action).grid(
                row=idx, column=2, sticky='w', padx=5)
            ttk.Label(scrollable, text=f'{entry_price:.2f}').grid(
                row=idx, column=3, sticky='w', padx=5)
            ttk.Label(scrollable, text=f'{sl_price:.2f}' if sl_price else '-').grid(
                row=idx, column=4, sticky='w', padx=5)
            ttk.Label(scrollable, text=f'{tp_price:.2f}' if tp_price else '-').grid(
                row=idx, column=5, sticky='w', padx=5)
            ttk.Label(scrollable, text=f'{volume:,.0f}').grid(
                row=idx, column=6, sticky='w', padx=5)

        btn_frame = ttk.Frame(frame)
        btn_frame.pack(fill=tk.X, pady=(10, 0))

        def add_selected():
            entries = []
            for info in vars_map.values():
                if info['var'].get():
                    side_map = {'BUY': 'LONG', 'SELL': 'SHORT'}
                    entries.append(DiaryEntry(
                        date=info['date'],
                        ticker=info['ticker'],
                        side=side_map.get(info['action'], info['action']),
                        entry_price=info['entry_price'],
                        sl_price=info['sl_price'],
                        tp_price=info['tp_price'],
                        volume=info['volume'],
                        qty=info['qty'],
                        max_hold=max_hold,
                        status='open'
                    ))
            if entries:
                self.diary_storage.add_entries(entries)
                if self.diary_ui:
                    self.diary_ui.refresh()
                win.destroy()
                self.scanner_ui.show_report(
                    f"Добавлено в дневник: {len(entries)} сделок.")
            else:
                from tkinter import messagebox as mb
                mb.showwarning('Нет выбора', 'Не выбрано ни одной сделки.')

        ttk.Button(btn_frame, text='Добавить выбранные',
                   command=add_selected).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text='Отмена',
                   command=win.destroy).pack(side=tk.LEFT, padx=5)

        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), 'units')

        canvas.bind_all('<MouseWheel>', _on_mousewheel)
        win.protocol('WM_DELETE_WINDOW', lambda: (
            canvas.unbind_all('<MouseWheel>'), win.destroy()
        ))

    def on_add_to_diary_analysis(self) -> None:
        """Добавить текущий сигнал из вкладки Анализа в дневник"""
        signal = self.app._last_signal
        params = self.app._last_params
        if not signal or signal.get('action') not in ('BUY', 'SELL'):
            mb.showinfo('В дневник', 'Нет активного сигнала BUY/SELL.')
            return
        if not params:
            mb.showinfo('В дневник', 'Нет параметров. Запустите backtest.')
            return

        ticker = self.app.get_selected_ticker()
        if not ticker:
            return

        confirm = mb.askyesno(
            'Подтверждение',
            f'Добавить {ticker} в торговый дневник?\n\n'
            f'Сигнал: {signal["action"]}\n'
            f'Цена входа: {signal.get("last_price", 0):.2f}\n'
            f'SL: {signal.get("sl_price", 0):.2f} | TP: {signal.get("tp_price", 0):.2f}'
        )
        if not confirm:
            return

        capital = params.get('capital', 1_000_000)
        risk_per_trade = params.get('risk_per_trade', 0.02)
        entry_price = signal.get('last_price') or signal.get('level', 0)
        sl_price = signal.get('sl_price', 0)
        tp_price = signal.get('tp_price', 0)
        qty = calc_position_qty(capital, risk_per_trade, entry_price, sl_price)
        volume = calc_position_volume(capital, risk_per_trade, entry_price, sl_price)
        side_map = {'BUY': 'LONG', 'SELL': 'SHORT'}

        entry = DiaryEntry(
            date=datetime.now().strftime('%Y-%m-%d %H:%M'),
            ticker=ticker,
            side=side_map.get(signal['action'], signal['action']),
            entry_price=entry_price,
            sl_price=sl_price,
            tp_price=tp_price,
            volume=volume,
            qty=qty,
            max_hold=params.get('max_hold', 20),
            status='open'
        )
        self.diary_storage.add_entries([entry])
        if self.diary_ui:
            self.diary_ui.refresh()
        self.notification_manager.on_position_opened(
            ticker, entry.side, entry_price)
        mb.showinfo('В дневник', f'{ticker} добавлен в дневник.')

    def on_add_to_pending(self) -> None:
        """Поставить лимитный ордер на вход по уровню сигнала."""
        signal = self.app._last_signal
        params = self.app._last_params
        if not signal or signal.get('action') not in ('BUY', 'SELL'):
            mb.showinfo('В ожидание', 'Нет активного сигнала BUY/SELL.')
            return
        if not params:
            mb.showinfo('В ожидание', 'Нет параметров. Запустите backtest.')
            return

        ticker = self.app.get_selected_ticker()
        if not ticker:
            return

        entry_price = signal.get('level') or signal.get('last_price', 0)
        sl_price = signal.get('sl_price', 0)
        tp_price = signal.get('tp_price', 0)

        confirm = mb.askyesno(
            'Ожидание сделки',
            f'Поставить лимитный ордер для {ticker}?\n\n'
            f'Сигнал: {signal["action"]}\n'
            f'Цена входа (уровень): {entry_price:.2f}\n'
            f'SL: {sl_price:.2f} | TP: {tp_price:.2f}\n\n'
            f'Монитор автоматически отследит касание цены\n'
            f'и переведёт ордер в дневник как открытую сделку.'
        )
        if not confirm:
            return

        capital = params.get('capital', 1_000_000)
        risk_per_trade = params.get('risk_per_trade', 0.02)
        side_map = {'BUY': 'LONG', 'SELL': 'SHORT'}
        side = side_map.get(signal['action'], signal['action'])
        qty = calc_position_qty(capital, risk_per_trade, entry_price, sl_price)
        volume = calc_position_volume(capital, risk_per_trade, entry_price, sl_price)

        trade = self.pending_storage.add_pending(
            ticker=ticker,
            side=side,
            entry_price=entry_price,
            sl_price=sl_price,
            tp_price=tp_price,
            qty=qty,
            volume=volume,
            capital=capital,
            risk_per_trade=risk_per_trade,
            max_hold=params.get('max_hold', 20),
            source='analysis',
        )
        self.pending_ui.update_trades(self.pending_storage.get_all())
        self.notification_manager.notify(
            'pending_triggered',
            f'Ордер установлен: {ticker}',
            f'{side} @ {entry_price:.2f} | SL {sl_price:.2f} | TP {tp_price:.2f}',
            icon='info',
        )
        mb.showinfo('В ожидание', f'Ордер для {ticker} установлен в очередь ожидания.')

    def _on_pending_remove(self, pending_id):
        """Удалить запись из ожидания или очистить сработавшие."""
        if pending_id == '__clear_triggered__':
            self.pending_storage.clear_triggered()
        else:
            for t in self.pending_storage.get_all():
                if t.pending_id == pending_id or t.created == pending_id:
                    self.pending_storage.remove_pending(t.pending_id)
                    break
        self.pending_ui.update_trades(self.pending_storage.get_all())

    def _on_pending_refresh(self):
        """Обновить таблицу ожидания."""
        self.pending_ui.update_trades(self.pending_storage.get_all())

    def _auto_pending_check_callback(self):
        """Фоновая задача: проверка срабатывания ожидающих ордеров."""
        if not self.pending_storage.get_active():
            return
        try:
            self._pending_do_check()
        except Exception as e:
            logging.warning(f'Pending check error: {e}')

    def _pending_do_check(self):
        """Выполнить проверку и обновить UI в главном потоке."""
        from datetime import datetime
        from pending.storage import check_entry_touch
        from diary.journal import DiaryEntry

        active = self.pending_storage.get_active()
        if not active:
            return

        triggered = []
        today_str = datetime.now().strftime('%Y-%m-%d')

        for trade in active:
            try:
                date_from = trade.created[:10]
                candles = self._fetch_candles_for_pending(trade.ticker, date_from, today_str)
                if isinstance(candles, str) or not isinstance(candles, list):
                    continue
                if len(candles) < 1:
                    continue
                if check_entry_touch(trade.entry_price, candles):
                    triggered.append(trade)
            except Exception as e:
                logging.debug(f'Pending check error for {trade.ticker}: {e}')

        for trade in triggered:
            now_str = datetime.now().strftime('%Y-%m-%d %H:%M')
            self.pending_storage.mark_triggered(trade.pending_id, now_str)

            entry = DiaryEntry(
                date=now_str,
                ticker=trade.ticker,
                side=trade.side,
                entry_price=trade.entry_price,
                sl_price=trade.sl_price,
                tp_price=trade.tp_price,
                volume=trade.volume,
                qty=trade.qty,
                status='open',
                max_hold=trade.max_hold,
            )
            self.diary_storage.add_entries([entry])

            self.notification_manager.notify(
                'pending_triggered',
                f'Ордер активирован: {trade.ticker}',
                f'{trade.side} @ {trade.entry_price:.2f} | SL {trade.sl_price:.2f} | TP {trade.tp_price:.2f}',
                icon='info',
            )

        if triggered:
            if self.diary_ui:
                self.root.after(0, lambda: self.diary_ui.refresh())
            self.root.after(0, lambda: self.pending_ui.update_trades(self.pending_storage.get_all()))

    def _fetch_candles_for_pending(self, ticker, date_from, date_to):
        """Получить свечи для проверки ожидающего ордера."""
        try:
            return self.get_stock_data(ticker, date_from, date_to)
        except Exception as e:
            logging.debug(f'Fetch candles for pending {ticker}: {e}')
            return None

    def on_show_ticker_settings(self) -> None:
        """Показать индивидуальные настройки для каждой бумаги"""
        path = os.path.join(app_dir(), 'results', 'ticker_settings.json')
        if not os.path.exists(path):
            mb.showinfo('Индивид. настройки', 'Нет сохранённых настроек.')
            return
        import json
        with open(path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        if not data:
            mb.showinfo('Индивид. настройки', 'Нет сохранённых настроек.')
            return

        from strategy.config import get_strategy_names
        name_map = dict(get_strategy_names())

        win = tk.Toplevel(self.root)
        win.title('Индивидуальные настройки бумаг')
        win.geometry('700x550')

        text_frame = ttk.Frame(win)
        text_frame.pack(fill=tk.BOTH, expand=1, padx=5, pady=(5, 0))

        text = tk.Text(text_frame, wrap=tk.NONE, font=('Consolas', 10))
        text.pack(side=tk.LEFT, fill=tk.BOTH, expand=1)
        scroll_y = tk.Scrollbar(text_frame, orient='vertical', command=text.yview)
        scroll_y.pack(side=tk.RIGHT, fill=tk.Y)
        text.configure(yscrollcommand=scroll_y.set)

        content = []
        for ticker in sorted(data):
            entry = data[ticker]
            sid = entry.get('strategy', '?')
            sname = name_map.get(sid, sid)
            content.append(f' {ticker}  ({sname})')
            for k, v in entry.get('params', {}).items():
                content.append(f'   {k}: {v}')
            content.append('')
        text.insert('1.0', '\n'.join(content))
        text.configure(state='disabled')

        btn_frame = ttk.Frame(win)
        btn_frame.pack(fill=tk.X, padx=5, pady=5)

        ttk.Button(btn_frame, text='Экспорт JSON',
                   command=lambda: self._export_ticker_settings(data, win)).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text='Импорт JSON',
                   command=lambda: self._import_ticker_settings(data, win, name_map)).pack(side=tk.LEFT, padx=2)

    def _export_ticker_settings(self, data, win):
        from tkinter import filedialog
        path = filedialog.asksaveasfilename(
            defaultextension='.json',
            filetypes=[('JSON', '*.json')],
            title='Экспорт настроек бумаг'
        )
        if not path:
            return
        try:
            import json
            with open(path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            mb.showinfo('Экспорт', f'Настройки сохранены:\n{path}')
        except Exception as e:
            mb.showerror('Ошибка', f'Не удалось экспортировать:\n{e}')

    def _import_ticker_settings(self, current_data, win, name_map):
        from tkinter import filedialog
        path = filedialog.askopenfilename(
            defaultextension='.json',
            filetypes=[('JSON', '*.json')],
            title='Импорт настроек бумаг'
        )
        if not path:
            return
        try:
            import json
            with open(path, 'r', encoding='utf-8') as f:
                imported = json.load(f)
            if not isinstance(imported, dict):
                mb.showerror('Ошибка', 'Неверный формат файла.')
                return
            merged = dict(current_data)
            merged.update(imported)
            settings_path = os.path.join(app_dir(), 'results', 'ticker_settings.json')
            with open(settings_path, 'w', encoding='utf-8') as f:
                json.dump(merged, f, ensure_ascii=False, indent=2)
            mb.showinfo('Импорт',
                        f'Импортировано настроек: {len(imported)} бумаг.\n'
                        'Закройте и откройте окно заново для просмотра.')
        except Exception as e:
            mb.showerror('Ошибка', f'Не удалось импортировать:\n{e}')

    def on_check_positions(self) -> None:
        """Проверить открытые позиции — закрыть по SL/TP"""
        if not self.diary_ui:
            return

        def task():
            try:
                updated = self.diary_storage.check_positions(self.get_stock_data)
                self.root.after(0, lambda: self._on_check_positions_done(updated))
            except Exception as e:
                self.root.after(0, lambda: self.diary_ui.refresh())
                logging.exception('Check positions error')

        self.diary_ui.refresh()
        t = threading.Thread(target=task, daemon=True)
        t.start()

    def _auto_check_positions(self):
        open_entries = self.diary_storage.get_open_entries()
        if not open_entries:
            return

        def task():
            try:
                updated = self.diary_storage.check_positions(self.get_stock_data)
                if updated:
                    self.root.after(0, lambda: self.diary_ui.refresh())
                    self.root.after(0, lambda: mb.showinfo(
                        'Автопроверка',
                        f'Закрыто по SL/TP/TIMEOUT: {updated} сделок.'))
            except Exception:
                logging.exception('Auto check positions error')

        t = threading.Thread(target=task, daemon=True)
        t.start()

    def _update_market_regime(self):
        from market.regime import MarketRegimeDetector, REGIME_CRISIS, REGIME_TRENDING_UP, REGIME_TRENDING_DOWN, REGIME_RANGING
        detector = MarketRegimeDetector()

        def on_regime(result):
            regime = result.get('regime', REGIME_RANGING)
            label = result.get('label', 'Флэт')
            confidence = result.get('confidence', 0)

            color_map = {
                REGIME_TRENDING_UP: '#00aa00',
                REGIME_TRENDING_DOWN: '#cc0000',
                REGIME_RANGING: '#888888',
                REGIME_CRISIS: '#ff0000',
            }
            color = color_map.get(regime, '#888888')

            try:
                self.regime_label.config(text=f"Рынок: {label} ({confidence:.0%})", foreground=color)
            except Exception:
                pass

            self._current_regime = result

            self.notification_manager.check_regime_change(regime)

            self._update_breadth()

        detector.detect_async(on_regime)

    def _update_breadth(self):
        tickers = self.sector_db.get_all_tickers()
        if not tickers:
            try:
                self.breadth_label.config(text="Breadth: —")
            except Exception:
                pass
            self._schedule_regime_update()
            return

        sample = tickers[:50]

        def task():
            from datetime import datetime, timedelta
            from market.breadth import BreadthIndicator

            end_date = datetime.now().strftime('%Y-%m-%d')
            start_date = (datetime.now() - timedelta(days=120)).strftime('%Y-%m-%d')

            ticker_closes = {}
            for ticker in sample:
                try:
                    candles = self.get_stock_data(ticker, start_date, end_date)
                    if isinstance(candles, list) and len(candles) > 50:
                        closes = [float(c[1]) for c in candles if c and len(c) >= 2]
                        if len(closes) > 50:
                            ticker_closes[ticker] = closes
                except Exception:
                    continue

            indicator = BreadthIndicator(ma_period=50)
            result = indicator.calculate(ticker_closes)
            return result

        def on_done(result):
            self._last_breadth = result
            try:
                if result:
                    pct = result['above_pct']
                    zone = result['zone']
                    zone_labels = {
                        'overbought': 'перекупл.',
                        'oversold': 'перепрод.',
                        'neutral': 'нейтр.',
                    }
                    zone_text = zone_labels.get(zone, zone)
                    color_map = {
                        'overbought': '#cc6600',
                        'oversold': '#0088ff',
                        'neutral': '#888888',
                    }
                    color = color_map.get(zone, '#888888')
                    self.breadth_label.config(
                        text=f"Breadth: {pct:.0f}% ({zone_text}, {result['above_count']}/{result['total_count']})",
                        foreground=color)
                else:
                    self.breadth_label.config(text="Breadth: —")
            except Exception:
                pass

            self._schedule_regime_update()

        import threading
        def run():
            r = task()
            try:
                self.root.after(0, on_done, r)
            except Exception:
                pass

        threading.Thread(target=run, daemon=True).start()

    def _on_regime_click(self, event=None):
        regime = getattr(self, '_current_regime', None)
        if not regime:
            return

        win = tk.Toplevel(self.root)
        win.title('Детали рыночного режима')
        win.geometry('550x500')
        win.resizable(True, True)

        main = ttk.Frame(win, padding=10)
        main.pack(fill=tk.BOTH, expand=1)

        label = regime.get('label', '—')
        confidence = regime.get('confidence', 0)
        details = regime.get('details', {})
        recommended = regime.get('recommended_strategies', [])

        info_text = f"Регим: {label} (уверенность {confidence:.0%})\n\n"
        if details:
            info_text += "Показатели IMOEX:\n"
            labels_map = {
                'adx': 'ADX',
                'plus_di': '+DI',
                'minus_di': '-DI',
                'ma': 'SMA(50)',
                'close': 'Цена',
                'recent_return': 'Доходность 20д (%)',
                'atr_ratio': 'ATR ratio',
            }
            for key, val in details.items():
                lbl = labels_map.get(key, key)
                info_text += f"  {lbl}: {val}\n"

        if recommended:
            info_text += f"\nРекомендуемые стратегии: {', '.join(recommended)}"

        text = tk.Text(main, wrap='word', font=('Consolas', 10), padx=10, pady=10)
        scroll = ttk.Scrollbar(main, orient='vertical', command=text.yview)
        text.configure(yscrollcommand=scroll.set)
        text.pack(side=tk.LEFT, fill=tk.BOTH, expand=1)
        scroll.pack(side=tk.RIGHT, fill=tk.Y)
        text.insert('1.0', info_text)
        text.configure(state='disabled')

    def _on_breadth_click(self, event=None):
        breadth = getattr(self, '_last_breadth', None)
        if not breadth:
            return

        win = tk.Toplevel(self.root)
        win.title('Детали Breadth')
        win.geometry('600x550')
        win.resizable(True, True)

        main = ttk.Frame(win, padding=10)
        main.pack(fill=tk.BOTH, expand=1)

        pct = breadth['above_pct']
        total = breadth['total_count']
        above = breadth['above_count']
        zone = breadth['zone']
        zone_labels = {'overbought': 'Перекупленность', 'oversold': 'Перепроданность', 'neutral': 'Нейтрально'}

        header = ttk.Label(main, text=f"Выше MA(50): {above}/{total} ({pct:.1f}%) — {zone_labels.get(zone, zone)}",
                           font=('', 11, 'bold'))
        header.pack(anchor='w', pady=(0, 5))

        nb = ttk.Notebook(main)
        nb.pack(fill=tk.BOTH, expand=1)

        above_frame = ttk.Frame(nb)
        below_frame = ttk.Frame(nb)
        nb.add(above_frame, text=f'Выше MA ({above})')
        nb.add(below_frame, text=f'Ниже MA ({total - above})')

        cols = ('ticker', 'close', 'ma', 'dist')
        headers = {'ticker': 'Тикер', 'close': 'Цена', 'ma': 'MA(50)', 'dist': 'От MA %'}
        widths = {'ticker': 80, 'close': 90, 'ma': 90, 'dist': 80}

        above_tree = ttk.Treeview(above_frame, columns=cols, show='headings', height=20)
        for col in cols:
            above_tree.heading(col, text=headers[col])
            above_tree.column(col, width=widths[col], anchor='center')
        above_scroll = ttk.Scrollbar(above_frame, orient='vertical', command=above_tree.yview)
        above_tree.configure(yscrollcommand=above_scroll.set)
        above_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=1)
        above_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        above_tree.tag_configure('positive', foreground='#00aa00')

        above_items = []
        for item in breadth.get('above_items', []):
            dist = (item['close'] - item['ma']) / item['ma'] * 100 if item['ma'] else 0
            above_items.append({
                'values': (item['ticker'], f"{item['close']:.2f}", f"{item['ma']:.2f}", f"+{dist:.1f}%"),
                'tags': ('positive',),
            })
        tree_batch_insert(above_tree, above_items, clear=False)

        below_tree = ttk.Treeview(below_frame, columns=cols, show='headings', height=20)
        for col in cols:
            below_tree.heading(col, text=headers[col])
            below_tree.column(col, width=widths[col], anchor='center')
        below_scroll = ttk.Scrollbar(below_frame, orient='vertical', command=below_tree.yview)
        below_tree.configure(yscrollcommand=below_scroll.set)
        below_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=1)
        below_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        below_tree.tag_configure('negative', foreground='#cc0000')

        below_items = []
        for item in breadth.get('below_items', []):
            dist = (item['ma'] - item['close']) / item['ma'] * 100 if item['ma'] else 0
            below_items.append({
                'values': (item['ticker'], f"{item['close']:.2f}", f"{item['ma']:.2f}", f"-{dist:.1f}%"),
                'tags': ('negative',),
            })
        tree_batch_insert(below_tree, below_items, clear=False)

    def _schedule_regime_update(self):
        try:
            self.root.after(1800000, self._update_market_regime)
        except Exception:
            pass

    def _on_check_positions_done(self, updated):
        self.diary_ui.refresh()
        import tkinter.messagebox as mb
        if updated:
            mb.showinfo('Проверка позиций',
                        f'Закрыто по SL/TP: {updated} сделок.')
        else:
            mb.showinfo('Проверка позиций',
                        'Открытые позиции без изменений.')

    def _on_position_refresh(self):
        if self.position_monitor is None:
            from monitoring.position_monitor import PositionMonitor
            self.position_monitor = PositionMonitor(
                self.root, self.diary_storage, self.get_stock_data,
                on_alerts=self._on_position_alerts,
                on_refresh=self._on_position_data,
            )
        self.position_monitor.near_distance_pct = self.position_ui.get_near_distance_pct()
        self.position_monitor.check_once()

    def _on_position_start(self):
        if self.position_monitor is None:
            from monitoring.position_monitor import PositionMonitor
            self.position_monitor = PositionMonitor(
                self.root, self.diary_storage, self.get_stock_data,
                on_alerts=self._on_position_alerts,
                on_refresh=self._on_position_data,
            )
        self.position_monitor.set_interval(self.position_ui.get_interval_sec())
        self.position_monitor.near_distance_pct = self.position_ui.get_near_distance_pct()
        self.position_monitor.start()
        self.position_ui.set_monitor_running(True)

    def _on_position_stop(self):
        if self.position_monitor:
            self.position_monitor.stop()
        self.position_ui.set_monitor_running(False)

    def _on_position_alerts(self, alerts):
        self.position_ui.update_alerts(alerts)
        self.notification_manager.check_alerts(alerts)

    def _on_position_data(self, positions):
        self.position_ui.update_positions(positions)
        if self.diary_ui:
            self.diary_ui.refresh()

    def _init_automation(self):
        if self.scheduler.get_autostart_monitor():
            self._on_position_start()

        for name in ['auto_scan', 'monitor_positions', 'refresh_watchlist', 'check_pending_trades']:
            task = self.scheduler.get_task(name)
            if task and task.enabled and not task.is_running:
                self.scheduler.start_task(name)

    def _auto_scan_callback(self):
        tickers = self.watchlist_storage.get_tickers()
        if not tickers:
            return 'Нет тикеров в вочлисте'

        from screening.scanner import Scanner
        from screening.reporter import generate_report
        from datetime import datetime, timedelta
        end = datetime.now().strftime('%Y-%m-%d')
        start = (datetime.now() - timedelta(days=120)).strftime('%Y-%m-%d')
        params = dict(self._last_scan_params) if self._last_scan_params else {
            'capital': 1_000_000,
            'risk_per_trade': 0.02,
            'atr_sl': 1.0,
            'atr_tp': 2.0,
            'min_hits': 5,
            'max_hold': 20,
            'commission': 0.0005,
        }
        scanner = Scanner(sector_db=self.sector_db, fetch_fn=self.get_stock_data)
        ticker_settings_path = os.path.join(app_dir(), 'results', 'ticker_settings.json')
        sectors = self.sector_db.get_all_sectors()
        results = scanner.scan(
            sectors=sectors,
            date_from=start,
            date_to=end,
            backtest_params=params,
            ticker_settings_path=ticker_settings_path,
        )
        new_signals = 0
        for r in results:
            sig = r.get('signal', {})
            action = sig.get('action', 'NONE')
            if action in ('BUY', 'SELL'):
                ticker = r.get('ticker', '')
                side = 'Лонг' if action == 'BUY' else 'Шорт'
                price = r.get('last_price')
                sl = sig.get('sl')
                tp = sig.get('tp')
                existing = self.signal_storage.get_signals(ticker=ticker, limit=5)
                recent = [s for s in existing if s['strategy'] == 'auto_scan' and s['side'] == side]
                if not recent:
                    self.signal_storage.add_signal(
                        ticker=ticker, side=side, price=price,
                        strategy='auto_scan', sl=sl, tp=tp,
                    )
                    self.notification_manager.on_signal_detected(ticker, side, 'auto_scan', price)
                    new_signals += 1
        self._last_scan_results = results
        try:
            self.root.after(0, lambda: self.signal_journal_ui.update_signals(
                self.signal_storage.get_signals(limit=200)))
            self.root.after(0, lambda: self.signal_journal_ui.update_strategies(
                self.signal_storage.get_strategies()))
            self.root.after(0, self._update_top_signals)
        except Exception:
            pass
        return f'{len(results)} тикеров, {new_signals} новых сигналов'

    def _auto_monitor_callback(self):
        if self.position_monitor is None:
            from monitoring.position_monitor import PositionMonitor
            self.position_monitor = PositionMonitor(
                self.root, self.diary_storage, self.get_stock_data,
                on_alerts=self._on_position_alerts,
                on_refresh=self._on_position_data,
            )
        self.position_monitor.check_once()
        open_entries = self.diary_storage.get_open_entries()
        if open_entries:
            for entry in open_entries:
                if hasattr(entry, 'current_dd_pct') and entry.current_dd_pct:
                    self.notification_manager.check_drawdown(entry.current_dd_pct)
        return f'{len(open_entries)} открытых позиций'

    def _auto_watchlist_callback(self):
        self._on_watchlist_refresh()
        tickers = self.watchlist_storage.get_tickers()
        return f'{len(tickers)} тикеров обновлено'

    def _on_automation_start_all(self):
        self._on_position_start()

    def _on_automation_stop_all(self):
        self._on_position_stop()

    def _update_top_signals(self):
        signals = self.signal_storage.get_signals(limit=5)
        self.app.update_top_signals(signals)

    def _on_correlation_matrix(self):
        tickers = self.watchlist_storage.get_tickers()
        if len(tickers) < 2:
            import tkinter.messagebox as mb
            mb.showwarning('Корреляция', 'Нужно минимум 2 тикера в избранном.')
            return

        self._watchlist_set_status('Загрузка данных...')

        def run():
            from datetime import datetime, timedelta
            from concurrent.futures import ThreadPoolExecutor, as_completed
            end = datetime.now().strftime('%Y-%m-%d')
            start = (datetime.now() - timedelta(days=60)).strftime('%Y-%m-%d')

            def fetch_one(t):
                try:
                    candles = self.get_stock_data(t, start, end)
                    if isinstance(candles, list) and len(candles) >= 5:
                        closes = [float(c[1]) for c in candles if c and len(c) >= 2]
                        if len(closes) >= 5:
                            return t, closes
                except Exception:
                    pass
                return t, None

            price_data = {}
            with ThreadPoolExecutor(max_workers=5) as pool:
                futures = {pool.submit(fetch_one, t): t for t in tickers}
                for f in as_completed(futures, timeout=120):
                    t, closes = f.result()
                    if closes is not None:
                        price_data[t] = closes

            if len(price_data) < 2:
                self.root.after(0, lambda: self._watchlist_set_status('Недостаточно данных'))
                return

            from analytics.correlation import calc_correlation_matrix
            result = calc_correlation_matrix(price_data)

            def show():
                from analytics.visualizations import plot_correlation_heatmap
                fig = plot_correlation_heatmap(result)
                if fig:
                    import matplotlib
                    matplotlib.use('TkAgg')
                    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
                    win = tk.Toplevel(self.root)
                    win.title('Корреляционная матрица')
                    win.geometry('700x600')
                    canvas = FigureCanvasTkAgg(fig, master=win)
                    canvas.draw()
                    canvas.get_tk_widget().pack(fill=tk.BOTH, expand=1)
                    NavigationToolbar2Tk(canvas, win)
                    self._watchlist_set_status(
                        f'{len(result["tickers"])} тикеров, {len(result["pairs"])} пар')
                else:
                    self._watchlist_set_status('Ошибка построения графика')

            self.root.after(0, show)

        import threading
        threading.Thread(target=run, daemon=True).start()

    def _on_alert_add(self, alert):
        self._update_top_signals()

    def _on_alert_remove(self, alert_id):
        pass

    def _on_alert_start(self):
        self.alert_monitor.start()
        self.alert_ui.set_monitor_running(True)

    def _on_alert_stop(self):
        self.alert_monitor.stop()
        self.alert_ui.set_monitor_running(False)

    def _on_alert_triggered(self, alert, current_price):
        cond_label = 'выше' if alert.condition == 'above' else 'ниже'
        self.notification_manager.notify(
            'price_alert',
            f'Алерт: {alert.ticker}',
            f'Цена {current_price:.2f} {cond_label} {alert.target_price:.2f}',
            icon='warning',
        )
        self.alert_ui.refresh()

    def _on_alert_refresh(self, alerts):
        self.alert_ui.refresh(alerts)

    def _on_news_refresh(self):
        self.news_ui.status_label.configure(text='Загрузка новостей...')
        self._on_news_analyze()

    def _on_news_analyze(self):
        self.news_ui.status_label.configure(text='Анализ новостей...')

        def run():
            try:
                from news.provider import load_ai_config
                config = load_ai_config()
                max_news = config.get('max_news', 50)
                results = self.news_analyzer.fetch_and_analyze(count=max_news)

                def show():
                    self.news_ui.update_news(results)
                    self.news_ui.status_label.configure(
                        text=f'{len(results)} новостей проанализировано')
                self.root.after(0, show)
            except Exception as e:
                import traceback
                traceback.print_exc()
                err_msg = str(e)[:80]
                self.root.after(0, lambda: self.news_ui.status_label.configure(
                    text=f'Ошибка: {err_msg}'))

        import threading
        threading.Thread(target=run, daemon=True).start()

    def _on_news_settings(self):
        old = self.news_ui._on_settings
        self.news_ui._on_settings = None
        self.news_ui._show_settings(on_saved=self.news_analyzer._rebuild_provider)
        self.news_ui._on_settings = old

    def _auto_news_callback(self):
        from news.provider import load_ai_config
        config = load_ai_config()
        max_news = config.get('max_news', 50)
        results = self.news_analyzer.fetch_and_analyze(count=max_news)
        positive = [r for r in results if r.get('sentiment') == 'positive']
        negative = [r for r in results if r.get('sentiment') == 'negative']
        high_impact = [r for r in results if r.get('impact', 0) >= 3]
        for r in high_impact:
            tickers = r.get('tickers', [])
            if isinstance(tickers, list):
                for t in tickers:
                    self.notification_manager.notify(
                        'news_alert',
                        f'AI: {t} {r.get("sentiment", "")}',
                        r.get('summary', '')[:200],
                        icon='info',
                    )
        try:
            self.root.after(0, lambda: self.news_ui.update_news(results))
        except Exception:
            pass
        return f'{len(results)} новостей, +{len(positive)}/-{len(negative)}, {len(high_impact)} сильных'

    def _on_review_refresh(self):
        entries = self.diary_storage.load()
        if not entries:
            import tkinter.messagebox as mb
            mb.showwarning('Обзор', 'Дневник пуст. Добавьте сделки.')
            return
        closed = [e for e in entries if e.status == 'closed' and e.pnl is not None]
        if not closed:
            import tkinter.messagebox as mb
            mb.showwarning('Обзор', 'Нет закрытых сделок для анализа.')
            return
        from diary.trade_review import compute_review
        capital = self.review_ui.get_capital()
        result = compute_review(entries, capital=capital)
        self.review_ui.update_review(result)

    def _on_perf_analyze(self):
        if self._last_trades is None or not self._last_trades:
            entries = self.diary_storage.load()
            if not entries:
                import tkinter.messagebox as mb
                mb.showwarning('Аналитика', 'Нет данных. Запустите backtest или добавьте сделки в дневник.')
                return
            closed = [e for e in entries if e.status == 'closed' and e.pnl is not None]
            if not closed:
                import tkinter.messagebox as mb
                mb.showwarning('Аналитика', 'Нет закрытых сделок.')
                return
            from diary.journal import asdict
            trades = [asdict(e) for e in closed]
            for t in trades:
                t['pnl_pct'] = (t['pnl'] / t.get('entry_price', 1)) * 100 if t.get('entry_price') else 0
                t['exit_reason'] = t.get('exit_reason', 'Вручную')
                t['sl_price'] = t.get('sl_price', 0)
                t['entry_price'] = t.get('entry_price', 0)
                t['qty'] = t.get('qty', 1)
            capital = self._last_capital
        else:
            trades = self._last_trades
            capital = self._last_capital

        benchmark_returns = None
        bench = self.perf_ui.get_benchmark()
        if bench == 'IMOEX':
            try:
                from market.regime import load_imoex_candles, candles_to_df
                candles = load_imoex_candles()
                if candles:
                    df = candles_to_df(candles)
                    if df is not None and len(df) > 10:
                        closes = df['Close'].astype(float)
                        benchmark_returns = closes.pct_change().dropna().tolist()
            except Exception:
                pass

        from analytics.performance import calc_advanced_metrics
        metrics = calc_advanced_metrics(trades, initial_capital=capital,
                                        benchmark_returns=benchmark_returns)
        self.perf_ui.update_analytics(metrics, trades=trades)

    def _on_pairs_scan(self):
        tickers = self.pairs_ui.get_tickers()
        if len(tickers) < 2:
            import tkinter.messagebox as mb
            mb.showwarning('Пары', 'Укажите минимум 2 тикера через запятую.')
            return

        self.pairs_ui.set_status('Загрузка данных...')
        params = self.pairs_ui.get_params()

        def run():
            price_data = {}
            from datetime import datetime, timedelta
            end_date = datetime.now().strftime('%Y-%m-%d')
            start_date = (datetime.now() - timedelta(days=365)).strftime('%Y-%m-%d')

            for ticker in tickers:
                candles = self.get_stock_data(ticker, start_date, end_date)
                if isinstance(candles, list) and len(candles) > 30:
                    closes = []
                    for c in candles:
                        if c is not None and len(c) >= 2:
                            closes.append(float(c[1]))
                    if len(closes) > 30:
                        price_data[ticker] = closes

            if len(price_data) < 2:
                self.root.after(0, lambda: self.pairs_ui.set_status(
                    'Недостаточно данных. Проверьте тикеры.'))
                return

            self.root.after(0, lambda: self.pairs_ui.set_status(
                f'Анализ {len(price_data)} тикеров...'))

            from pairs.cointegration import find_cointegrated_pairs, format_pairs_report
            pairs = find_cointegrated_pairs(
                price_data,
                max_pairs=params['max_pairs'],
            )

            report = format_pairs_report(pairs)

            self.root.after(0, lambda: self._on_pairs_scan_complete(pairs, report))

        import threading
        threading.Thread(target=run, daemon=True).start()

    def _on_pairs_scan_complete(self, pairs, report):
        self.pairs_ui.update_pairs(pairs)
        self.pairs_ui.update_result(report)
        self.pairs_ui.set_status(f'Найдено пар: {len(pairs)}')

    def _on_pairs_backtest(self):
        pair = self.pairs_ui.get_selected_pair()
        if not pair:
            import tkinter.messagebox as mb
            mb.showwarning('Пары', 'Выберите пару из таблицы.')
            return

        params = self.pairs_ui.get_params()
        ticker_y = pair['ticker_y']
        ticker_x = pair['ticker_x']

        self.pairs_ui.set_status(f'Загрузка {ticker_y}, {ticker_x}...')

        def run():
            from datetime import datetime, timedelta
            end_date = datetime.now().strftime('%Y-%m-%d')
            start_date = (datetime.now() - timedelta(days=365)).strftime('%Y-%m-%d')

            candles_y = self.get_stock_data(ticker_y, start_date, end_date)
            candles_x = self.get_stock_data(ticker_x, start_date, end_date)

            if not isinstance(candles_y, list) or not isinstance(candles_x, list):
                self.root.after(0, lambda: self.pairs_ui.set_status('Ошибка загрузки данных'))
                return

            closes_y = [float(c[1]) for c in candles_y if c and len(c) >= 2]
            closes_x = [float(c[1]) for c in candles_x if c and len(c) >= 2]
            dates_y = [str(c[6])[:10] for c in candles_y if c and len(c) > 6]

            min_len = min(len(closes_y), len(closes_x), len(dates_y))
            if min_len < params['lookback'] + 10:
                self.root.after(0, lambda: self.pairs_ui.set_status('Слишком мало данных'))
                return

            closes_y = closes_y[-min_len:]
            closes_x = closes_x[-min_len:]
            dates = dates_y[-min_len:]

            from pairs.strategy import run_pairs_backtest, format_pairs_backtest_report
            result = run_pairs_backtest(
                price_y=closes_y,
                price_x=closes_x,
                dates=dates,
                hedge_ratio=pair['hedge_ratio'],
                capital=self._last_capital,
                entry_z=params['entry_z'],
                exit_z=params['exit_z'],
                stop_z=params['stop_z'],
                max_hold=params['max_hold'],
                lookback=params['lookback'],
            )

            report = format_pairs_backtest_report(result, ticker_y, ticker_x)

            self.root.after(0, lambda: self._on_pairs_bt_complete(report, result, pair))

        import threading
        threading.Thread(target=run, daemon=True).start()

    def _on_pairs_bt_complete(self, report, result, pair):
        self.pairs_ui.update_result(report)
        self.pairs_ui.set_status(
            f"{pair['ticker_y']}/{pair['ticker_x']}: "
            f"{result['total_trades']} сделок, "
            f"P&L={result['total_pnl']:+,.0f}")

        try:
            import matplotlib
            matplotlib.use('TkAgg')
            import matplotlib.pyplot as plt
            from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

            if result.get('trades') and pair.get('spread'):
                spread = pair['spread']
                fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(10, 4))

                ax1.plot(spread, linewidth=0.8, color='#2196F3')
                mu = pair['spread_mean']
                sigma = pair['spread_std']
                ax1.axhline(y=mu, color='gray', linestyle='--', linewidth=0.5)
                ax1.axhline(y=mu + 2 * sigma, color='red', linestyle=':', linewidth=0.5)
                ax1.axhline(y=mu - 2 * sigma, color='red', linestyle=':', linewidth=0.5)
                ax1.set_title(f"Спред {pair['ticker_y']}/{pair['ticker_x']}")
                ax1.set_ylabel('Спред')

                if result.get('equity_curve'):
                    ax2.plot(result['equity_curve'], linewidth=1, color='#44aa44')
                    ax2.set_title('Equity парной стратегии')
                    ax2.set_ylabel('Капитал')

                fig.tight_layout()
                plt.show()
        except Exception:
            pass

    def _init_new_tabs(self):
        self._on_watchlist_refresh()
        self.signal_journal_ui.update_strategies(self.signal_storage.get_strategies())
        self.signal_journal_ui.update_signals(self.signal_storage.get_signals(limit=200))
        self._update_top_signals()

    def _on_watchlist_add(self, ticker):
        self.watchlist_storage.add(ticker)
        self._on_watchlist_refresh()

    def _watchlist_set_status(self, text):
        if hasattr(self, 'watchlist_ui') and self.watchlist_ui is not None:
            try:
                self.watchlist_ui.set_status(text)
            except Exception:
                pass

    def _on_watchlist_remove(self, ticker):
        self.watchlist_storage.remove(ticker)
        self._on_watchlist_refresh()

    def _on_watchlist_refresh(self):
        tickers = self.watchlist_storage.get_tickers()
        if not tickers:
            if hasattr(self, 'watchlist_ui') and self.watchlist_ui is not None:
                self.watchlist_ui.update_watchlist([])
                self.watchlist_ui.set_status('Список пуст')
            return

        self._watchlist_set_status(f'Обновление {len(tickers)} тикеров...')

        def task():
            from datetime import datetime, timedelta
            from fundamental.data import fetch_all_tqbr_fundamentals, fetch_dividends, calc_dividend_yield
            end_date = datetime.now().strftime('%Y-%m-%d')
            start_date = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
            fund_data = {}
            try:
                fund_data = fetch_all_tqbr_fundamentals()
            except Exception:
                pass
            items = []
            for ticker in tickers:
                try:
                    candles = self.get_stock_data(ticker, start_date, end_date)
                    price = None
                    prev_close = None
                    volume = None
                    change_pct = 0
                    if isinstance(candles, list) and len(candles) >= 2:
                        last = candles[-1]
                        prev = candles[-2]
                        price = float(last[1]) if last and len(last) >= 2 else None
                        prev_close = float(prev[1]) if prev and len(prev) >= 2 else None
                        volume = int(float(last[4])) if last and len(last) >= 5 else None
                        change_pct = ((price / prev_close) - 1) * 100 if price and prev_close else 0

                    sector = self.sector_db.get_sector(ticker) or ''
                    fd = fund_data.get(ticker, {})
                    market_cap = fd.get('market_cap')
                    div_yield = None
                    if price:
                        try:
                            divs = fetch_dividends(ticker)
                            div_yield = calc_dividend_yield(divs, price) if divs else None
                        except Exception:
                            pass

                    items.append({
                        'ticker': ticker,
                        'price': price,
                        'change_pct': change_pct,
                        'volume': volume,
                        'sector': sector,
                        'market_cap': market_cap,
                        'div_yield': div_yield,
                    })
                except Exception:
                    items.append({'ticker': ticker, 'price': None, 'change_pct': 0,
                                  'volume': None, 'sector': self.sector_db.get_sector(ticker) or '',
                                  'market_cap': None, 'div_yield': None})
            return items

        import threading
        def run():
            result = task()
            try:
                self.root.after(0, lambda: self._on_watchlist_done(result))
            except Exception:
                pass

        threading.Thread(target=run, daemon=True).start()

    def _on_watchlist_done(self, items):
        if hasattr(self, 'watchlist_ui') and self.watchlist_ui is not None:
            self.watchlist_ui.update_watchlist(items)
            self.watchlist_ui.set_status(f'Обновлено: {len(items)} тикеров')

    def _on_watchlist_select(self, ticker):
        if ticker:
            self.ticker_var.set(ticker)

    def _on_dividend_calendar(self):
        tickers = self.watchlist_storage.get_tickers()
        if not tickers:
            self._watchlist_set_status('Список пуст')
            return
        self._watchlist_set_status('Загрузка дивидендов...')

        def run():
            try:
                from fundamental.data import fetch_dividends
                from datetime import datetime
                lines = ['=' * 60, '  ДИВИДЕНДНЫЙ КАЛЕНДАРЬ', '=' * 60, '']
                today = datetime.now()
                any_found = False
                for ticker in sorted(tickers):
                    try:
                        divs = fetch_dividends(ticker)
                    except Exception:
                        continue
                    if not divs:
                        continue
                    for d in divs:
                        date_str = d.get('date', '') or d.get('registryclosedate', '')
                        val = d.get('value') or d.get('amount', 0)
                        cur = d.get('currencyid', 'RUB')
                        if not date_str:
                            continue
                        try:
                            div_date = datetime.strptime(date_str[:10], '%Y-%m-%d')
                        except (ValueError, TypeError):
                            continue
                        if (div_date - today).days >= -30:
                            any_found = True
                            lines.append(f'  {ticker:8s}  {date_str[:10]}  {float(val):>8.2f} {cur}')
                if not any_found:
                    lines.append('  Нет предстоящих дивидендов')
                lines.append('')
                report = '\n'.join(lines)
                self.root.after(0, lambda: self._show_popup('Дивидендный календарь', report))
                self.root.after(0, lambda: self._watchlist_set_status('Готово'))
            except Exception as e:
                self.root.after(0, lambda: self._watchlist_set_status(f'Ошибка: {e}'))

        import threading
        threading.Thread(target=run, daemon=True).start()

    def _on_sector_scan(self):
        period = self.sector_rotation_ui.get_period()
        self.sector_rotation_ui.set_status('Сканирование секторов...')

        sector_tickers_map = {}
        for sector in self.sector_db.get_all_sectors():
            tickers = self.sector_db.get_tickers([sector])
            if tickers:
                sector_tickers_map[sector] = tickers

        from market.sector_rotation import calc_sector_rotation_async

        def on_done(results):
            self.sector_rotation_ui.update_sectors(results)
            self.sector_rotation_ui.set_status(f'Найдено секторов: {len(results)}')

        calc_sector_rotation_async(on_done, sector_tickers_map, period)

    def _on_sector_compare(self):
        self.sector_rotation_ui.set_status('Загрузка фундаментал. данных...')
        sector_map = self.sector_db.get_ticker_to_sector_map()

        def run():
            try:
                from fundamental.data import fetch_all_tqbr_fundamentals, fetch_dividends, calc_dividend_yield
                from fundamental.filters import compare_sector
                all_fund = fetch_all_tqbr_fundamentals()
                for ticker, td in all_fund.items():
                    price = td.get('prev_price')
                    if price:
                        try:
                            divs = fetch_dividends(ticker)
                            td['div_yield'] = calc_dividend_yield(divs, price) if divs else None
                            td['dividends'] = divs
                            td['div_count'] = len(divs)
                        except Exception:
                            pass
                for ticker in list(all_fund.keys()):
                    if ticker in sector_map:
                        all_fund[ticker]['sector'] = sector_map[ticker]
                result = compare_sector(all_fund)
                lines = ['=' * 60, '  СРАВНЕНИЕ СЕКТОРОВ (фундаментал)', '=' * 60, '']
                for sec, d in sorted(result.items(), key=lambda x: x[1].get('avg_div_yield', 0) or 0, reverse=True):
                    avg_pe = d.get('avg_pe')
                    avg_pb = d.get('avg_pb')
                    avg_div = d.get('avg_div_yield')
                    avg_cap = d.get('avg_market_cap')
                    n = d.get('count', 0)
                    pe_str = f'{avg_pe:.1f}' if avg_pe else '—'
                    pb_str = f'{avg_pb:.2f}' if avg_pb else '—'
                    div_str = f'{avg_div:.1f}%' if avg_div else '—'
                    if avg_cap:
                        if avg_cap >= 1e12:
                            cap_str = f'{avg_cap / 1e12:.1f}T'
                        elif avg_cap >= 1e9:
                            cap_str = f'{avg_cap / 1e9:.0f}M'
                        else:
                            cap_str = f'{avg_cap / 1e6:.0f}M'
                    else:
                        cap_str = '—'
                    lines.append(f'  {sec:20s}  P/E:{pe_str:>7s}  P/B:{pb_str:>7s}  Див:{div_str:>7s}  Кап:{cap_str:>7s}  ({n} тик.)')
                lines.append('')
                report = '\n'.join(lines)
                self.root.after(0, lambda: self.sector_rotation_ui.set_status('Готово'))
                self.root.after(0, lambda: self._show_popup('Сравнение секторов', report))
            except Exception as e:
                self.root.after(0, lambda: self.sector_rotation_ui.set_status(f'Ошибка: {e}'))

        import threading
        threading.Thread(target=run, daemon=True).start()

    def _show_popup(self, title, text):
        popup = tk.Toplevel(self.root)
        popup.title(title)
        popup.geometry('750x500')
        popup.grab_set()
        txt = tk.Text(popup, wrap=tk.WORD, font=('Consolas', 10))
        sb = ttk.Scrollbar(popup, orient='vertical', command=txt.yview)
        txt.configure(yscrollcommand=sb.set)
        txt.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        sb.pack(side=tk.RIGHT, fill=tk.Y)
        txt.insert('1.0', text)
        txt.configure(state='disabled')

    def _on_pos_calc(self):
        params = self.pos_calc_ui.get_params()
        from analytics.position_calc import calc_position, format_position_report
        result = calc_position(
            entry_price=params['entry_price'],
            sl_price=params['sl_price'],
            capital=params['capital'],
            risk_pct=params['risk_pct'],
            tp_price=params['tp_price'] or None,
        )
        if result:
            result['ticker'] = params['ticker']
        report = format_position_report(result)
        self.pos_calc_ui.show_result(report)

    def _get_current_price(self, ticker):
        try:
            from datetime import datetime, timedelta
            end_date = datetime.now().strftime('%Y-%m-%d')
            start_date = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')
            candles = self.get_stock_data(ticker, start_date, end_date)
            if isinstance(candles, list) and len(candles) >= 1:
                last = candles[-1]
                if last and len(last) >= 2:
                    return float(last[1])
        except Exception:
            pass
        return None

    def _on_signal_filter(self):
        filters = self.signal_journal_ui.get_filters()
        signals = self.signal_storage.get_signals(
            ticker=filters['ticker'] or None,
            strategy=filters['strategy'],
            side=filters['side'],
        )
        self.signal_journal_ui.update_strategies(self.signal_storage.get_strategies())
        self.signal_journal_ui.update_signals(signals)

    def _on_signal_export(self):
        from tkinter import filedialog
        path = filedialog.asksaveasfilename(
            defaultextension='.csv',
            filetypes=[('CSV', '*.csv')],
            title='Экспорт сигналов в CSV',
        )
        if not path:
            return
        filters = self.signal_journal_ui.get_filters()
        count = self.signal_storage.export_csv(
            path,
            ticker=filters['ticker'] or None,
            strategy=filters['strategy'],
            side=filters['side'],
        )
        import tkinter.messagebox as mb
        mb.showinfo('Экспорт', f'Экспортировано {count} сигналов в\n{path}')

    def on_show_analysis(self) -> None:
        """Показать окно анализа сделок с equity curve"""
        entries = self.diary_storage.load()
        closed = [e for e in entries if e.status == 'closed' and e.pnl is not None]
        if not closed:
            import tkinter.messagebox as mb
            mb.showwarning('Анализ', 'Нет закрытых сделок для анализа.')
            return

        import matplotlib.pyplot as plt
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

        win = tk.Toplevel(self.root)
        win.title("Анализ сделок")
        win.geometry("900x600")

        capital = self._last_capital
        equity = [capital]
        dates = [closed[0].date[:10]]
        running = capital

        cumulative_pnl = 0
        wins = 0
        losses = 0

        for e in closed:
            running += (e.pnl or 0)
            equity.append(running)
            dates.append(e.exit_date[:10] if e.exit_date else e.date[:10])
            cumulative_pnl += (e.pnl or 0)
            if (e.pnl or 0) > 0:
                wins += 1
            else:
                losses += 1

        fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(10, 6), gridspec_kw={'height_ratios': [3, 1]})

        ax1.plot(dates, equity, 'b-', linewidth=1.5, label='Капитал')
        ax1.fill_between(dates, capital, equity, alpha=0.1, color='blue')
        ax1.axhline(y=capital, color='gray', linestyle='--', linewidth=0.8)
        ax1.set_ylabel('Капитал (₽)')
        ax1.set_title('Кривая капитала')
        ax1.legend()
        ax1.tick_params(axis='x', rotation=45)
        ax1.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'{x:,.0f}'))

        total_pnl = sum(e.pnl for e in closed)
        total_trades = len(closed)
        win_rate = (wins / total_trades * 100) if total_trades else 0

        stats_text = (
            f'Сделок: {total_trades} | '
            f'Win Rate: {win_rate:.1f}% | '
            f'Прибыль: {total_pnl:+,.0f} ₽ | '
            f'Текущий капитал: {running:,.0f} ₽'
        )

        ax2.axis('off')
        ax2.text(0.5, 0.5, stats_text, ha='center', va='center',
                 fontsize=12, fontfamily='monospace',
                 bbox=dict(boxstyle='round,pad=0.8', facecolor='lightgray'))

        fig.tight_layout()

        canvas = FigureCanvasTkAgg(fig, master=win)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=1)

        from matplotlib.backends.backend_tkagg import NavigationToolbar2Tk
        toolbar = NavigationToolbar2Tk(canvas, win)
        toolbar.update()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=1)

    def on_show_legend(self) -> None:
        """Показать окно с легендой сигналов"""
        from screening.reporter import get_legend_text
        legend_win = tk.Toplevel(self.root)
        legend_win.title("Легенда сигналов")
        legend_win.geometry("550x500")
        text_w = tk.Text(legend_win, wrap=tk.WORD, font=('Consolas', 10))
        text_w.pack(fill=tk.BOTH, expand=1, padx=10, pady=10)
        text_w.insert(tk.END, get_legend_text())
        text_w.config(state=tk.DISABLED)
        _add_copy_menu(text_w)


def main():
    """Точка входа приложения"""
    import tempfile
    _err = os.path.join(tempfile.gettempdir(), 'CreateStrategy_error.log')
    try:
        os.makedirs(os.path.join(app_dir(), 'results'), exist_ok=True)
        root = tk.Tk()
        app = CreateStrategyApp(root)
        try:
            root.mainloop()
        except KeyboardInterrupt:
            app._on_close()
    except Exception:
        import traceback
        with open(_err, 'w', encoding='utf-8') as f:
            f.write(traceback.format_exc())
        raise


if __name__ == "__main__":
    main()
